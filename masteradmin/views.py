from django.shortcuts import render, redirect
from django.views.generic import TemplateView
from .models import * 
from django.utils import timezone
from datetime import timedelta
from customersupport.models import *
from django.core.mail import send_mail
from django.conf import settings
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.views import View
import json, logging, logging.handlers, random, re, zipfile
from agents.models import *
from app.models import *
from django.contrib import messages
from openpyxl import Workbook
from django.utils.translation import gettext_lazy as _
from django.db import IntegrityError
from employee.models import *  
from django.contrib.auth.hashers import check_password, make_password
from User.models import *
from User.models import Feedbacks
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.views.decorators.http import require_http_methods
from django.http import JsonResponse
from decimal import Decimal
from django.views.decorators.csrf import csrf_protect
from django.db.models import Count
from django.db.models import Q
from django.db.models import Sum, Count, Avg
from django.db.models.functions import TruncDay
from beesuggest.models import ProductDetails, BeesuggestAgreement

# register = template.Library()

# @register.filter(name='compress_html')
# def compress_html(value):
#     """Minifies HTML by removing newlines, multiple spaces, and comments."""
#     value = re.sub(r'\n+', '', value)  # Remove newlines
#     value = re.sub(r'\s+', ' ', value)  # Replace multiple spaces with a single space
#     value = re.sub(r'<!--.*?-->', '', value)  # Remove HTML comments
#     return value.strip()


# Configure logging
logger = logging.getLogger(__name__)
file_handler = logging.handlers.RotatingFileHandler(
    'masteradmin.log',
    maxBytes=5*1024*1024,  # 5MB
    backupCount=3,
    encoding='utf-8'
)
console_handler = logging.StreamHandler()
log_format = logging.Formatter('%(asctime)s [%(levelname)s] %(message)s - %(pathname)s:%(lineno)d')
file_handler.setFormatter(log_format)
console_handler.setFormatter(log_format)
logger.addHandler(file_handler)
logger.addHandler(console_handler)
logger.setLevel(logging.INFO)

# def masteradmin_login(request):
#     logger.info("masteradmin_login view called")
#     if request.method == 'POST':
#         action = request.POST.get('action')
#         logger.debug(f"POST request received with action: {action}")
        
        # if action == 'verify_email':
        #     email = request.POST.get('email')
        #     if not email:
        #         logger.warning("Email field is empty")
        #         return JsonResponse({'status': 'error', 'message': 'Email is required'})
                
            # Generate OTP and send email
            # email_otp = ''.join([str(random.randint(0, 9)) for _ in range(6)])
            # request.session['email_otp'] = email_otp
            # request.session['email'] = email
            # logger.debug(f"Generated email OTP for {email}")
            
            # try:
            #     send_mail(
            #         'Master Admin Login OTP',
            #         f'Your OTP for email verification is: {email_otp}',
            #         settings.DEFAULT_FROM_EMAIL,
            #         [email],
            #         fail_silently=False,
            #     )
            #     logger.info(f"OTP sent successfully to {email}")
            #     return JsonResponse({'status': 'success', 'message': 'OTP sent to email'})
            # except Exception as e:
            #     logger.error(f"Failed to send email OTP: {str(e)}")
            #     return JsonResponse({'status': 'error', 'message': str(e)})
                
        # elif action == 'verify_email_otp':
        #     entered_otp = request.POST.get('email_otp')
        #     stored_otp = request.session.get('email_otp')
        #     logger.debug(f"Verifying email OTP - entered: {entered_otp}, stored: {stored_otp}")
            
        #     if entered_otp == stored_otp:
        #         logger.info("Email OTP verified successfully")
        #         return JsonResponse({'status': 'success', 'message': 'Email verified'})
        #     logger.warning("Invalid email OTP entered")
        #     return JsonResponse({'status': 'error', 'message': 'Invalid OTP'})
            
        # elif action == 'verify_mobile':
        #     mobile = request.POST.get('mobile')
        #     if not mobile:
        #         logger.warning("Mobile number field is empty") 
        #         return JsonResponse({'status': 'error', 'message': 'Mobile number is required'})
                
        #     # Generate OTP and send SMS
        #     mobile_otp = ''.join([str(random.randint(0, 9)) for _ in range(6)])
        #     request.session['mobile_otp'] = mobile_otp
        #     request.session['mobile'] = mobile
        #     logger.debug(f"Generated mobile OTP for {mobile}")
            
        #     # Send SMS logic here
        #     # For now just return success
        #     logger.info(f"Mobile OTP generated for {mobile}")
        #     return JsonResponse({'status': 'success', 'message': 'OTP sent to mobile'})
            
        # elif action == 'verify_mobile_otp':
        #     entered_otp = request.POST.get('mobile_otp') 
        #     stored_otp = request.session.get('mobile_otp')
        #     logger.debug(f"Verifying mobile OTP - entered: {entered_otp}, stored: {stored_otp}")
            
        #     if entered_otp == stored_otp:
        #         logger.info("Mobile OTP verified successfully")
        #         return JsonResponse({'status': 'success', 'message': 'Mobile verified'})
        #     logger.warning("Invalid mobile OTP entered")
        #     return JsonResponse({'status': 'error', 'message': 'Invalid OTP'})
            
    #     elif action == 'login':
    #         password = request.POST.get('password')
    #         email = request.session.get('email')
    #         mobile = request.session.get('mobile')
    #         logger.debug(f"Login attempt for email: {email}, mobile: {mobile}")
            
    #         if not all([email, mobile, password]):
    #             logger.warning("Missing required login fields")
    #             return JsonResponse({'status': 'error', 'message': 'All fields required'})
                
    #         # Verify credentials and login
    #         try:
    #             # Check if user exists with provided email
    #             user = MasterAdmin.objects.get(email=email)
    #             logger.debug(f"Found user with email {email}")
                
    #             # Verify password
    #             if not user.check_password(password):
    #                 logger.warning(f"Invalid password for user {email}")
    #                 return JsonResponse({'status': 'error', 'message': 'Invalid credentials'})
                
    #             # Check if user is a master admin
    #             if not user.is_superuser:
    #                 logger.warning(f"Unauthorized access attempt by {email}")
    #                 return JsonResponse({'status': 'error', 'message': 'Unauthorized access'})
                
    #             # Login the user
    #             masteradmin_login(request, user)
    #             logger.info(f"User {email} logged in successfully")
                
    #         except MasterAdmin.DoesNotExist:
    #             logger.warning(f"Invalid login attempt for non-existent user {email}")
    #             return JsonResponse({'status': 'error', 'message': 'Invalid credentials'})
            
    #         return JsonResponse({
    #             'status': 'success',
    #             'message': 'Login successful',
    #             'redirect': '/masteradmin/dashboard/'
    #         })
            
    # logger.debug("Rendering login page")
    # return render(request, 'masteradmin/login.html')

def MasterAdminLogin(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        password = request.POST.get('password')
        
        if not email:
            messages.error(request, 'Please provide both email and password')
            return render(request, 'masteradmin/login.html')
            
        try:
            user = MasterAdmin.objects.get(email=email)
            password = check_password(password, user.password)
            if password:
                request.session['masteradmin_id'] = str(user.id)
                return redirect('masteradmin_dashboard')
            else:
                messages.error(request, 'Invalid password')
        except MasterAdmin.DoesNotExist:
            messages.error(request, 'Invalid credentials')
            
    return render(request, 'masteradmin/login.html')


# def Signup(request):
#     logger.info("Signup view called")
#     if request.method == 'POST':
#         action = request.POST.get('action')
#         logger.info(f"POST request received with action: {action}")
        
#         if action == 'send_email_otp':
#             email = request.POST.get('email')
#             logger.info(f"Sending email OTP for: {email}")
            
#             if not email:
#                 logger.error("Email field is empty")
#                 return JsonResponse({'status': 'error', 'message': 'Email is required'})
                
#             # Check if email already exists
#             if MasterAdmin.objects.filter(email=email).exists():
#                 logger.warning(f"Email {email} already registered")
#                 return JsonResponse({'status': 'error', 'message': 'Email already registered'})
                
#             # Generate and store email OTP
#             email_otp = ''.join(random.choices('0123456789', k=6))
#             request.session['email_otp'] = email_otp
#             request.session['email'] = email
#             logger.info(f"Generated email OTP for {email}")
            
#             # Send email with OTP
#             try:
#                 subject = '1matrix Master Admin Signup - Email Verification'
#                 message = f'Your OTP for email verification is: {email_otp}'
#                 from_email = settings.EMAIL_HOST_USER
#                 recipient_list = [email]
#                 send_mail(subject, message, from_email, recipient_list)
#                 logger.info(f"Email OTP sent successfully to {email}")
                
#                 return JsonResponse({'status': 'success', 'message': 'OTP sent to email'})
#             except Exception as e:
#                 logger.error(f"Failed to send email OTP: {str(e)}")
#                 return JsonResponse({'status': 'error', 'message': 'Failed to send email OTP'})
            
#         elif action == 'verify_email_otp':
#             entered_otp = request.POST.get('email_otp')
#             stored_otp = request.session.get('email_otp')
#             logger.info("Verifying email OTP")
            
#             if not stored_otp:
#                 logger.error("No stored OTP found in session")
#                 return JsonResponse({'status': 'error', 'message': 'OTP expired. Please request new OTP'})
                
#             if entered_otp == stored_otp:
#                 request.session['email_verified'] = True
#                 logger.info("Email OTP verified successfully")
#                 return JsonResponse({'status': 'success', 'message': 'Email verified'})
#             logger.warning("Invalid email OTP entered")
#             return JsonResponse({'status': 'error', 'message': 'Invalid OTP'})
            
#         elif action == 'send_mobile_otp':
#             mobile = request.POST.get('mobile')
#             logger.info(f"Sending mobile OTP for: {mobile}")
            
#             if not mobile:
#                 logger.error("Mobile number field is empty")
#                 return JsonResponse({'status': 'error', 'message': 'Mobile number is required'})
                
#             # Validate mobile number format
#             if not re.match(r'^\+?1?\d{9,15}$', mobile):
#                 logger.error(f"Invalid mobile number format: {mobile}")
#                 return JsonResponse({'status': 'error', 'message': 'Invalid mobile number format'})
                
#             # Generate and store mobile OTP
#             mobile_otp = ''.join(random.choices('0123456789', k=6))
#             request.session['mobile_otp'] = mobile_otp
#             request.session['mobile'] = mobile
#             logger.info(f"Generated mobile OTP for {mobile}")
            
#             # Send SMS with OTP using your SMS service
#             try:
#                 # Send OTP via Twilio
#                 from twilio.rest import Client
#                 client = Client(settings.TWILIO_ACCOUNT_SID, settings.TWILIO_AUTH_TOKEN)
#                 try:
#                     message = client.messages.create(
#                         body=f'Your 1matrix verification code is: {mobile_otp}',
#                         from_=settings.TWILIO_PHONE_NUMBER,
#                         to=mobile
#                     )
#                     logger.info(f"Successfully sent Twilio SMS to {mobile} with message SID: {message.sid}")
#                 except Exception as twilio_error:
#                     logger.error(f"Twilio SMS failed: {str(twilio_error)}")
#                     raise Exception(f"Failed to send SMS via Twilio: {str(twilio_error)}")
                
#                 # For now just print the OTP
#                 logger.info(f"Mobile OTP: {mobile_otp}")
#                 return JsonResponse({'status': 'success', 'message': 'OTP sent to mobile'})
#             except Exception as e:
#                 logger.error(f"Failed to send mobile OTP: {str(e)}")
#                 return JsonResponse({'status': 'error', 'message': 'Failed to send mobile OTP'})
            
#         elif action == 'verify_mobile_otp':
#             entered_otp = request.POST.get('mobile_otp')
#             stored_otp = request.session.get('mobile_otp')
#             logger.info("Verifying mobile OTP")
            
#             if not stored_otp:
#                 logger.error("No stored mobile OTP found in session")
#                 return JsonResponse({'status': 'error', 'message': 'OTP expired. Please request new OTP'})
                
#             if entered_otp == stored_otp:
#                 request.session['mobile_verified'] = True
#                 logger.info("Mobile OTP verified successfully")
#                 return JsonResponse({'status': 'success', 'message': 'Mobile verified'})
#             logger.warning("Invalid mobile OTP entered")
#             return JsonResponse({'status': 'error', 'message': 'Invalid OTP'})
            
#         elif action == 'signup':
#             password = request.POST.get('password')
#             email = request.session.get('email')
#             mobile = request.session.get('mobile')
#             logger.info(f"Processing signup for email: {email}")
            
#             # Verify both email and mobile are verified
#             if not request.session.get('email_verified'):
#                 logger.error("Email not verified for signup")
#                 return JsonResponse({'status': 'error', 'message': 'Email not verified'})
                
#             if not request.session.get('mobile_verified'):
#                 logger.error("Mobile not verified for signup")
#                 return JsonResponse({'status': 'error', 'message': 'Mobile not verified'})
            
#             if not all([email, mobile, password]):
#                 logger.error("Missing required fields for signup")
#                 return JsonResponse({'status': 'error', 'message': 'All fields required'})
                
#             try:
#                 # Create new master admin user
#                 user = MasterAdmin.objects.create_superuser(
#                     email=email,
#                     mobile=mobile,
#                     password=password
#                 )
#                 logger.info(f"Created new master admin user: {email}")
                
#                 # Clear all session data
#                 request.session.flush()
#                 logger.info("Cleared session data")
                
#                 # Login the user
#                 masteradmin_login(request, user)
#                 logger.info(f"Logged in new master admin: {email}")
                
#                 return JsonResponse({
#                     'status': 'success', 
#                     'message': 'Signup successful',
#                     'redirect': '/masteradmin/dashboard/'
#                 })
                
#             except Exception as e:
#                 logger.error(f"Error during signup: {str(e)}")
#                 return JsonResponse({'status': 'error', 'message': str(e)})
                
#     logger.info("Rendering signup page")
#     return render(request, 'masteradmin/signup.html').

def Signup(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        password = request.POST.get('password')
        mobile = request.POST.get('mobile')
        if MasterAdmin.objects.filter(email=email).exists():
            messages.error(request, 'Email already registered')
            return redirect('/masteradmin/signup/')
        else:
            password2 = make_password(password)
            user = MasterAdmin(
                email=email,
                password=password2,
                phone_number=mobile,
                created_at=timezone.now(),
            )

            user.save()
            messages.success(request, 'Signup successful')
            return redirect('/masteradmin/login/')
    return render(request, 'masteradmin/signup.html')
            

def logout(request):
    logger.info("Logging out master admin")
    request.session.clear()
    return redirect('masteradmin_login')

@method_decorator(csrf_exempt, name='dispatch')
class AddAgentView(View):
    def post(self, request):
        logger.info("AddAgentView.post() called")
        try:
            if not request.body:
                logger.error("Empty request body received")
                return JsonResponse({"status": "error", "message": "Empty request body"}, status=400)

            # Parse JSON data
            try:
                data = json.loads(request.body.decode('utf-8'))
                logger.info(f"Request body: {request.body}")
                logger.info(f"Parsed data: {data}")
            except json.JSONDecodeError as e:
                logger.error(f"JSON decode error: {str(e)}")
                # Try to get form data instead
                data = request.POST
                if not data:
                    return JsonResponse({"status": "error", "message": "Invalid JSON format and no form data found"}, status=400)
                logger.info(f"Using form data instead: {data}")
            
            agent_email = data.get('email')
            department_id = data.get('department')

            if not all([agent_email, department_id]):
                logger.error(f"Missing required fields - email: {bool(agent_email)}, department: {bool(department_id)}")
                return JsonResponse({"status": "error", "message": "Missing required fields"}, status=400)

            try:
                department_instance = Policy.objects.get(id=department_id)
            except Policy.DoesNotExist:
                logger.error(f"Department with id {department_id} not found")
                return JsonResponse({"status": "error", "message": "Invalid department"}, status=400)

            logger.info(f"Extracted fields - email: {agent_email}, department: {department_id}")

            # Check if email already exists
            if AgentUser.objects.filter(email=agent_email).exists():
                logger.error(f"Agent with email {agent_email} already exists")
                return JsonResponse({"status": "error", "message": "Email already registered"}, status=400)

            # Generate new passcode
            last_agent = AgentUser.objects.order_by('-agent_passcode').first()
            if last_agent and last_agent.agent_passcode:
                new_passcode = int(last_agent.agent_passcode) + 1
            else:
                new_passcode = 7000

            # Create new agent with only required fields, rest as null
            try:
                agent = AgentUser.objects.create(
                    email=agent_email,
                    agent_passcode=str(new_passcode),
                    department=department_instance,
                    photo=None,
                    name=None,
                    password=None,
                    phone_number=None,
                    pan_number=None,
                    aadhar_number=None,
                    address=None,
                    dob=None,
                    qualification=None,
                    qualification_file=None,
                    bank_account_holder_name=None,
                    bank_account_number=None,
                    bank_name=None,
                    branch_name=None,
                    bank_ifsc_code=None,
                    cancelled_cheque_file=None,
                    offer_letter_file=None,
                    bank_statement_file=None,
                    increment_letter_file=None,
                    pay_slip_file=None,
                    experience_letter_file=None,
                    leave_letter_file=None,
                    addhar_card_file=None
                )

                # Generate form link with passcode
                form_link = f"https://{request.get_host()}/agents/register/{new_passcode}/"
                
                # email with form link
                email_subject = "Complete Your Agent Registration"
                email_message = f"""
                Hello,

                Please complete your agent registration by clicking the link below:
                {form_link}

                Your passcode: {new_passcode}

                Best regards,
                1Matrix
                """

                # Add proper email sending with error handling and retries
                max_retries = 3
                retry_count = 0
                email_sent = False

                while retry_count < max_retries and not email_sent:
                    try:
                        send_mail(
                            subject=email_subject,
                            message=email_message,
                            from_email=settings.EMAIL_HOST_USER,
                            recipient_list=[agent_email],
                            fail_silently=False,
                            auth_user=settings.EMAIL_HOST_USER,
                            auth_password=settings.EMAIL_HOST_PASSWORD
                        )
                        email_sent = True
                        logger.info(f"Registration email sent successfully to {agent_email}")
                    except Exception as e:
                        retry_count += 1
                        logger.error(f"Email sending attempt {retry_count} failed: {str(e)}")
                        if retry_count == max_retries:
                            # Delete the agent if email sending completely fails
                            agent.delete()
                            logger.error("Max retries reached for email sending. Agent creation rolled back.")
                            return JsonResponse({
                                "status": "error", 
                                "message": "Failed to send registration email. Please try again."
                            }, status=500)

                logger.info(f"Agent created successfully with email: {agent_email} and passcode: {new_passcode}")
                return JsonResponse({
                    "status": "success",
                    "message": "Agent created successfully",
                    "passcode": new_passcode
                })

            except Exception as e:
                logger.error(f"Error creating agent: {str(e)}")
                return JsonResponse({"status": "error", "message": "Error creating agent"}, status=500)

        except Exception as e:
            logger.error(f"Unexpected error: {str(e)}")
            return JsonResponse({"status": "error", "message": str(e)}, status=500)

class MasterAdminView(TemplateView):
    template_name = 'masteradmin/index.html'

    def get_context_data(self, **kwargs):
        logger.info("Getting context data for MasterAdminView")
        context = super().get_context_data(**kwargs)
        context['subscription_plans'] = Subscription.objects.all()
        context['user_policies'] = UserPolicy.objects.all()
        context['pinned_notes'] = QuickNotes.objects.filter(is_pinned=True)
        # Get counts for pending approvals
        pending_agents = AgentUser.objects.filter(name__isnull=False, is_approved=False).count()
        pending_employees = Employee.objects.filter(name__isnull=False, is_approved=False).count() 
        pending_support = SupportUser.objects.filter(name__isnull=False, is_approved=False).count()
        
        # Add to context
        context['total_agents'] = pending_agents + pending_employees + pending_support
        context['total_meetings'] = Meeting.objects.filter(is_active=True).count()
        context['departments'] = Policy.objects.all()
        context['support_departments'] = SupportDepartment.objects.all()
        
        # Add user creation form context
        context['user_types'] = [
            ('agent', 'Agent'),
            ('employee', 'Employee'),
            ('support', 'Support Staff')
        ]
        context['departments'] = Policy.objects.all()
        context['support_departments'] = SupportDepartment.objects.all()
        
        logger.debug("Returning context data")
        return context
    
class CreateUserView(View):
    def post(self, request):
        try:
            # Extract form data
            name = request.POST.get('name')
            email = request.POST.get('email')
            phone = request.POST.get('phone')
            gstin = request.POST.get('gstin')
            address = request.POST.get('address')
            subscription_id = request.POST.get('subscription')
            user_policy_id = request.POST.get('user_policy')
            discount = request.POST.get('discount', '0')
            agent_passcode = request.POST.get('agent_passcode')
            payment_mode = request.POST.get('payment_mode')
            upi_id = request.POST.get('upi_id')

            # Basic validation
            if not all([name, email, phone, address, subscription_id, user_policy_id, payment_mode, upi_id]):
                messages.error(request, 'Please fill all required fields')
                return redirect('masteradmin_dashboard')

            # Check for existing user
            if User.objects.filter(email=email).exists():
                messages.error(request, 'Email already registered')
                return redirect('masteradmin_dashboard')

            if User.objects.filter(phone=phone).exists():
                messages.error(request, 'Phone number already registered')
                return redirect('masteradmin_dashboard')

            # Get related objects
            try:
                subscription = Subscription.objects.get(id=subscription_id)
                user_policy = UserPolicy.objects.get(id=user_policy_id)
            except (Subscription.DoesNotExist, UserPolicy.DoesNotExist):
                messages.error(request, 'Invalid subscription or user policy')
                return redirect('masteradmin_dashboard')

            # Handle agent if passcode provided
            agent = None
            if agent_passcode:
                try:
                    agent = AgentUser.objects.get(agent_passcode=agent_passcode)
                except AgentUser.DoesNotExist:
                    messages.error(request, 'Invalid agent passcode')
                    return redirect('masteradmin_dashboard')

            # Create user
            user = User.objects.create(
                name=name,
                email=email,
                phone=phone,
                gst_number=gstin,
                address=address,
                subscription_plan=subscription,
                agent=agent,
                discount=float(discount) if discount else 0,
                user_policy=user_policy.description,
                is_active=True,
                upi_id=upi_id,
                last_payment_date=timezone.now(),
                last_payment_amount=subscription.price_monthly,
                last_payment_status='Pending',
                last_payment_mode=payment_mode
            )

            # Send welcome email with link
            try:
                subject = 'Welcome to Our Platform'
                message = f'''Hi {name},

Welcome to our platform! Your account has been created successfully.

You can access your account here: {request.build_absolute_uri('/login/')}

Your account details:
Email: {email}
Subscription Plan: {subscription.name}

If you have any questions, please don't hesitate to contact us.

Best regards,
The Team'''

                send_mail(
                    subject,
                    message,
                    settings.DEFAULT_FROM_EMAIL,
                    [email],
                    fail_silently=False,
                )
                logger.info(f"Welcome email sent successfully to user {email}")
            except Exception as e:
                logger.error(f"Failed to send welcome email to {email}: {str(e)}")
                # Continue execution even if email fails

            messages.success(request, 'User created successfully')
            logger.info(f"User created successfully: {user.email}")
            return redirect('masteradmin_dashboard')

        except Exception as e:
            logger.error(f"Error creating user: {str(e)}")
            messages.error(request, f'Error creating user: {str(e)}')
            return redirect('masteradmin_dashboard')

    def get(self, request):
        messages.error(request, 'Invalid request method')
        return redirect('masteradmin_dashboard')
class AgentsView(TemplateView):
    template_name = 'masteradmin/agents.html'

    def get_context_data(self, **kwargs):
        logger.info("Getting context data for AgentsView")
        context = super().get_context_data(**kwargs)
        context['agents'] = AgentUser.objects.all()
        context['new_registrations'] = AgentUser.objects.filter(is_approved=True).count()
        context['active_agents'] = AgentUser.objects.filter(is_approved=True, is_active=True).count()
        context['agents_total'] = AgentUser.objects.count()
        context['agents_active'] = AgentUser.objects.filter(is_approved=True, is_active=True)
        context['meetings'] = Meeting.objects.all()
        context['total_meetings'] = Meeting.objects.count()
        context['active_meetings'] = Meeting.objects.filter(is_active=True).count()
        context['completed_meetings'] = Meeting.objects.filter(is_completed=True).count()
        logger.debug(f"Meeting counts - total: {context['total_meetings']}, active: {context['active_meetings']}, completed: {context['completed_meetings']}")
        context['total_agents'] = AgentUser.objects.filter(name__isnull=False, is_approved=False).count() + Employee.objects.filter(name__isnull=False, is_approved=False).count() + SupportUser.objects.filter(name__isnull=False, is_approved=False).count()
        context['pending_agents'] = AgentUser.objects.filter(is_approved=False).count()
        context['departments'] = Policy.objects.all()
        context['total_meetings'] = Meeting.objects.filter(is_active=True).count()
        context['departments'] = Policy.objects.all()
        context['support_departments'] = SupportDepartment.objects.all()
        logger.debug(f"Agent counts - new: {context['new_registrations']}, active: {context['active_agents']}, total: {context['total_agents']}, pending: {context['pending_agents']}")
        return context

class SubscriptionView(TemplateView):
    template_name = 'masteradmin/subscription.html'

    def get_context_data(self, **kwargs):
        logger.info("Getting context data for SubscriptionView")
        context = super().get_context_data(**kwargs)
        context['subscriptions'] = Subscription.objects.all()
        context['total_subscriptions'] = Subscription.objects.count()
        # context['django_apps'] = 
        # logger.debug(f"Found {len(context['django_apps'])} custom Django apps")
        context['apps'] = Apps.objects.all()
        context['total_agents'] = AgentUser.objects.filter(name__isnull=False, is_approved=False).count() + Employee.objects.filter(name__isnull=False, is_approved=False).count() + SupportUser.objects.filter(name__isnull=False, is_approved=False).count()
        context['total_meetings'] = Meeting.objects.filter(is_active=True).count()
        context['departments'] = Policy.objects.all()
        context['support_departments'] = SupportDepartment.objects.all()
        logger.debug(f"Found {context['total_subscriptions']} subscriptions")
        return context

class AppsView(TemplateView):
    template_name = "masteradmin/app.html"

    def get_context_data(self, **kwargs):
        logger.info("Getting context data for AppsView")
        context = super().get_context_data(**kwargs)
        context['total_agents'] = AgentUser.objects.filter(name__isnull=False, is_approved=False).count() + Employee.objects.filter(name__isnull=False, is_approved=False).count() + SupportUser.objects.filter(name__isnull=False, is_approved=False).count()
        context['total_meetings'] = Meeting.objects.filter(is_active=True).count()
        context['departments'] = Policy.objects.all()
        context['support_departments'] = SupportDepartment.objects.all()
        # Include apps in the context
        context['apps'] = Apps.objects.all().order_by('-created_at')
        logger.debug("Returning context data")
        return context

@method_decorator(csrf_exempt, name='dispatch')
class CreateAppView(View):
    """View to create a new app."""
    def post(self, request, *args, **kwargs):
        name = request.POST.get('name')
        description = request.POST.get('description')
        url_keyword = request.POST.get('url_keyword', '').strip()
        is_active = request.POST.get('is_active') == 'true'

        if not name or not description:
            messages.error(request, 'Name and description are required.')
            return redirect('/masteradmin/apps/')
            
        try:
            Apps.objects.create(
                name=name,
                description=description,
                url_keyword=url_keyword if url_keyword else None,
                is_active=is_active
            )
            messages.success(request, f'App "{name}" created successfully.')
        except IntegrityError:
            messages.error(request, f'An app with the name "{name}" or URL keyword "{url_keyword}" already exists.')
        except Exception as e:
            messages.error(request, f'An error occurred: {e}')
            
        return redirect('/masteradmin/apps/')


@method_decorator(csrf_exempt, name='dispatch')
class UpdateAppView(View):
    """View to update an existing app."""
    def post(self, request, *args, **kwargs):
        app_id = request.POST.get('app_id')
        name = request.POST.get('name')
        description = request.POST.get('description')
        url_keyword = request.POST.get('url_keyword', '').strip()
        is_active = request.POST.get('is_active') == 'true'
        
        if not all([app_id, name, description]):
            messages.error(request, 'Missing required fields.')
            return redirect('/masteradmin/apps/')
            
        try:
            app = Apps.objects.get(id=app_id)
            app.name = name
            app.description = description
            app.url_keyword = url_keyword if url_keyword else None
            app.is_active = is_active
            app.save()
            messages.success(request, f'App "{name}" updated successfully.')
        except Apps.DoesNotExist:
            messages.error(request, 'App not found.')
        except IntegrityError:
            messages.error(request, f'An app with the name "{name}" or URL keyword "{url_keyword}" already exists.')
        except Exception as e:
            messages.error(request, f'An error occurred: {e}')
            
        return redirect('/masteradmin/apps/')

@method_decorator(csrf_exempt, name='dispatch')
class DeleteAppView(View):
    def post(self, request, *args, **kwargs):
        logger.info("Processing app deletion request")
        try:
            app_id = request.POST.get('app_id')
            
            if not app_id:
                logger.error("App ID is required")
                messages.error(request, 'App ID is required')
                return redirect('/masteradmin/apps/')
            
            app = Apps.objects.get(id=app_id)
            app_name = app.name
            
            # Check if app is used in any subscriptions
            subscriptions_using_app = Subscription.objects.filter(features=app)
            if subscriptions_using_app.exists():
                logger.warning(f"Cannot delete app {app_name} as it is used in {subscriptions_using_app.count()} subscription(s)")
                messages.error(request, f'Cannot delete "{app_name}" as it is being used in {subscriptions_using_app.count()} subscription plan(s).')
                return redirect('/masteradmin/apps/')
                
            app.delete()
            logger.info(f"Deleted app: {app_name}")
            messages.success(request, f'App "{app_name}" has been deleted successfully.')
            
        except Apps.DoesNotExist:
            logger.error(f"App with ID {app_id} not found")
            messages.error(request, f'App not found.')
        except Exception as e:
            logger.error(f"Error deleting app: {str(e)}")
            messages.error(request, f'Error deleting app: {str(e)}')
            
        return redirect('apps')

@method_decorator(csrf_exempt, name='dispatch')
class ToggleAppStatusView(View):
    def post(self, request, *args, **kwargs):
        app_id = request.POST.get('app_id')
        try:
            app = Apps.objects.get(id=app_id)
            # Toggle the is_temporarily_disabled status
            app.is_temporarily_disabled = not app.is_temporarily_disabled
            app.save()
            
            status_text = "disabled" if app.is_temporarily_disabled else "enabled"
            messages.success(request, f'The app "{app.name}" has been temporarily {status_text}.')
            logger.info(f'App "{app.name}" (ID: {app_id}) has been temporarily {status_text}.')
            
        except Apps.DoesNotExist:
            messages.error(request, 'The specified app could not be found.')
            logger.error(f'Attempted to toggle status for non-existent app with ID: {app_id}')
        except Exception as e:
            messages.error(request, f'An unexpected error occurred: {e}')
            logger.critical(f'An unexpected error occurred while toggling app status for ID {app_id}: {e}', exc_info=True)
            
        return redirect('apps')

class GetAppView(View):
    def get(self, request, app_id, *args, **kwargs):
        try:
            app = Apps.objects.get(id=app_id)
            data = {
                'id': app.id,
                'name': app.name,
                'description': app.description,
                'url_keyword': app.url_keyword,
                'is_active': app.is_active,
            }
            return JsonResponse({'status': 'success', 'app': data})
        except Apps.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'App not found'}, status=404)

class FeedbacksView(TemplateView):
    template_name = "masteradmin/feedbacks.html"

    def get_context_data(self, **kwargs):
        logger.info("Getting context data for FeedbacksView")
        context = super().get_context_data(**kwargs)
        context['total_agents'] = AgentUser.objects.filter(name__isnull=False, is_approved=False).count() + Employee.objects.filter(name__isnull=False, is_approved=False).count() + SupportUser.objects.filter(name__isnull=False, is_approved=False).count()
        context['total_meetings'] = Meeting.objects.filter(is_active=True).count()
        context['departments'] = Policy.objects.all()
        context['support_departments'] = SupportDepartment.objects.all()
        context['feedbacks'] = Feedbacks.objects.all()
        # Calculate feedback metrics
        all_feedbacks = Feedbacks.objects.count()
        
        # Positive feedbacks (4 and 5 stars)
        positive_feedbacks = Feedbacks.objects.filter(rating__in=[4, 5]).count()
        
        # Negative feedbacks (2 and 3 stars)
        negative_feedbacks = Feedbacks.objects.filter(rating__in=[2, 3]).count()
        
        # One star feedbacks
        one_star_feedbacks = Feedbacks.objects.filter(rating=1).count()
        
        # Calculate percentages (avoid division by zero)
        positive_percentage = (positive_feedbacks / all_feedbacks * 100) if all_feedbacks > 0 else 0
        negative_percentage = (negative_feedbacks / all_feedbacks * 100) if all_feedbacks > 0 else 0
        one_star_percentage = (one_star_feedbacks / all_feedbacks * 100) if all_feedbacks > 0 else 0
        
        # Add metrics to context
        context['all_feedbacks'] = all_feedbacks
        context['positive_feedbacks'] = positive_feedbacks
        context['negative_feedbacks'] = negative_feedbacks
        context['one_star_feedbacks'] = one_star_feedbacks
        context['positive_percentage'] = round(positive_percentage, 1)
        context['negative_percentage'] = round(negative_percentage, 1)
        context['one_star_percentage'] = round(one_star_percentage, 1)
        
        logger.debug(f"Feedback metrics - all: {all_feedbacks}, positive: {positive_feedbacks}, negative: {negative_feedbacks}, one star: {one_star_feedbacks}")
        logger.debug("Returning context data")
        return context
    

class UsersView(TemplateView):
    template_name = "masteradmin/users.html"

    def get_context_data(self, **kwargs):
        logger.info("Getting context data for UsersView")
        context = super().get_context_data(**kwargs)
        context['total_agents'] = AgentUser.objects.filter(name__isnull=False, is_approved=False).count() + Employee.objects.filter(name__isnull=False, is_approved=False).count() + SupportUser.objects.filter(name__isnull=False, is_approved=False).count()
        context['total_meetings'] = Meeting.objects.filter(is_active=True).count()
        context['departments'] = Policy.objects.all()
        context['support_departments'] = SupportDepartment.objects.all()
        context['users'] = User.objects.all()
        logger.debug("Returning context data")
        return context

class CustomerSupportView(TemplateView):
    template_name = "masteradmin/customer_support.html"

    def get_context_data(self, **kwargs):
        logger.info("Getting context data for CustomerSupportView")
        context = super().get_context_data(**kwargs)
        context['total_agents'] = AgentUser.objects.filter(name__isnull=False, is_approved=False).count() + Employee.objects.filter(name__isnull=False, is_approved=False).count() + SupportUser.objects.filter(name__isnull=False, is_approved=False).count()
        context['total_meetings'] = Meeting.objects.filter(is_active=True).count()
        context['departments'] = Policy.objects.all()
        context['support_departments'] = SupportDepartment.objects.all()
        logger.debug("Returning context data")
        return context
    
class SettingView(TemplateView):
    template_name = "masteradmin/settings.html"

    def get_context_data(self, **kwargs):
        logger.info("Getting context data for SettingView")
        context = super().get_context_data(**kwargs)
        context['total_agents'] = AgentUser.objects.filter(name__isnull=False, is_approved=False).count() + Employee.objects.filter(name__isnull=False, is_approved=False).count() + SupportUser.objects.filter(name__isnull=False, is_approved=False).count()
        context['total_meetings'] = Meeting.objects.filter(is_active=True).count()
        context['departments'] = Policy.objects.all()
        context['support_departments'] = SupportDepartment.objects.all()
        logger.debug("Returning context data")
        return context
    
class SalesView(TemplateView):
    template_name = "masteradmin/sales.html"

    def get_context_data(self, **kwargs):
        logger.info("Getting context data for SalesView")
        context = super().get_context_data(**kwargs)
        context['total_agents'] = AgentUser.objects.filter(name__isnull=False, is_approved=False).count() + Employee.objects.filter(name__isnull=False, is_approved=False).count() + SupportUser.objects.filter(name__isnull=False, is_approved=False).count()
        context['total_meetings'] = Meeting.objects.filter(is_active=True).count()
        context['departments'] = Policy.objects.all()
        context['support_departments'] = SupportDepartment.objects.all()
        logger.debug("Returning context data")
        return context

class CreateSubscriptionView(TemplateView):
    template_name = "masteradmin/create_subscription.html"

    def get_context_data(self, **kwargs):
        logger.info("Getting context data for CreateSubscriptionView")
        context = super().get_context_data(**kwargs)
        context['total_agents'] = AgentUser.objects.filter(name__isnull=False, is_approved=False).count() + Employee.objects.filter(name__isnull=False, is_approved=False).count() + SupportUser.objects.filter(name__isnull=False, is_approved=False).count()
        context['total_meetings'] = Meeting.objects.filter(is_active=True).count()
        context['departments'] = Policy.objects.all()
        context['support_departments'] = SupportDepartment.objects.all()
        context['apps'] = Apps.objects.all()
        logger.debug("Returning context data")
        return context
    
    def post(self, request, *args, **kwargs):
        logger.info("Processing subscription creation request")
        try:
            data = request.POST
            
            # Validate required fields
            name = data.get('name')
            if not name:
                logger.error("Name field is missing")
                messages.error(request, 'Name is required')
                return redirect('create_subscription')
            
            # Create subscription with all model fields
            subscription = Subscription.objects.create(
                name=name,
                plan_type=data.get('plan_type'),
                price_monthly=float(data.get('price_monthly', 0)),
                users=int(data.get('users', 1)), 
                max_users=int(data.get('max_users', 4)),
                validity_days=int(data.get('validity_days', 30)),
                discount=float(data.get('discount', 0)) if data.get('discount') else None,
                additional_user_cost=float(data.get('additional_user_cost', 1000)),
                is_active=True,
                paused=False
            )
            logger.info(f"Created subscription: {subscription.name}")

            # Process app features and their limits
            selected_features = request.POST.getlist('features[]')
            if selected_features:
                logger.debug(f"Processing {len(selected_features)} selected apps")
                
                # Filter valid feature IDs and add to subscription
                valid_features = Apps.objects.filter(id__in=selected_features)
                subscription.features.set(valid_features)
                
                # Process app limits for each selected app
                for app_id in selected_features:
                    limit_type = request.POST.get(f'limit_type_{app_id}')
                    
                    if limit_type == 'Limited':
                        # Process limited type
                        credits = request.POST.get(f'app_credits_{app_id}')
                        daily_limit = request.POST.get(f'app_daily_limit_{app_id}')
                        price = request.POST.get(f'app_price_{app_id}')
                        
                        # Convert to appropriate types
                        credits = int(credits) if credits else None
                        daily_limit = int(daily_limit) if daily_limit else None
                        price = float(price) if price else 0
                        
                        AppSubscriptionLimit.objects.create(
                            subscription=subscription,
                            app_id=app_id,
                            limit_type='Limited',
                            credits=credits,
                            daily_limit=daily_limit,
                            price=price
                        )
                        logger.debug(f"Added limited app settings for app {app_id}")
                    
                    elif limit_type == 'Unlimited':
                        # Process unlimited type
                        price = request.POST.get(f'app_unlimited_price_{app_id}')
                        price = float(price) if price else 0
                        
                        AppSubscriptionLimit.objects.create(
                            subscription=subscription,
                            app_id=app_id,
                            limit_type='Unlimited',
                            credits=None,
                            daily_limit=None,
                            price=price
                        )
                        logger.debug(f"Added unlimited app settings for app {app_id}")
            
            # Process coupons
            coupon_names = request.POST.getlist('coupon_name[]')
            coupon_codes = request.POST.getlist('coupon_code[]')
            coupon_discount_types = request.POST.getlist('coupon_discount_type[]')
            coupon_rates = request.POST.getlist('coupon_rate[]')
            coupon_number_of_uses = request.POST.getlist('coupon_number_of_uses[]')
            coupon_expiry_dates = request.POST.getlist('coupon_expiry_date[]')
            
            # Create coupons
            for i in range(len(coupon_names)):
                if i < len(coupon_codes) and coupon_names[i] and coupon_codes[i]:
                    discount_type = coupon_discount_types[i] if i < len(coupon_discount_types) else 'Flat'
                    rate = float(coupon_rates[i]) if i < len(coupon_rates) and coupon_rates[i] else 0
                    uses = int(coupon_number_of_uses[i]) if i < len(coupon_number_of_uses) and coupon_number_of_uses[i] else 1
                    
                    # Handle expiry date
                    expiry_date = None
                    if i < len(coupon_expiry_dates) and coupon_expiry_dates[i]:
                        try:
                            expiry_date = coupon_expiry_dates[i]
                        except Exception as e:
                            logger.error(f"Error parsing expiry date: {str(e)}")
                    
                    # Create the coupon
                    coupon = Coupons.objects.create(
                        name=coupon_names[i],
                        code=coupon_codes[i],
                        discount_type=discount_type,
                        rate=rate,
                        number_of_uses=uses,
                        expiry_date=expiry_date
                    )
                    
                    # Associate coupon with subscription
                    subscription.coupons.add(coupon)
                    logger.debug(f"Created and added coupon: {coupon.name}")

            messages.success(request, 'Subscription created successfully')
            return redirect('subscription')
            
        except ValueError as e:
            logger.error(f"Invalid numeric values: {str(e)}")
            messages.error(request, f'Invalid numeric values provided: {str(e)}')
            return redirect('create_subscription')
        except Exception as e:
            logger.error(f"Error creating subscription: {str(e)}")
            messages.error(request, f'Error creating subscription: {str(e)}')
            return redirect('create_subscription')

# HTML Template Usage:
# Form: <form method="POST" action="{% url 'delete_subscription' subscription.id %}">
# Link: {% url 'delete_subscription' subscription.id %}
# Note: subscription.id should be a UUID string
class DeleteSubscriptionView(TemplateView):
    template_name = 'masteradmin/delete_subscription.html'

    def get(self, request, *args, **kwargs):
        context = super().get_context_data(**kwargs)
        context['total_agents'] = AgentUser.objects.filter(name__isnull=False, is_approved=False).count() + Employee.objects.filter(name__isnull=False, is_approved=False).count() + SupportUser.objects.filter(name__isnull=False, is_approved=False).count()
        context['total_meetings'] = Meeting.objects.filter(is_active=True).count()
        context['apps'] = Apps.objects.all()
        context['departments'] = Policy.objects.all()
        context['support_departments'] = SupportDepartment.objects.all()
        context['total_subscriptions'] = Subscription.objects.count()
        logger.info("Processing subscription deletion GET request")
        try:
            # Get all subscriptions to display in list
            subscriptions = Subscription.objects.all()
            context = {
                'subscriptions': subscriptions
            }
            logger.debug(f"Found {subscriptions.count()} subscriptions")
            return render(request, self.template_name, context)
        except Exception as e:
            logger.error(f"Error loading subscriptions: {str(e)}")
            messages.error(request, f'Error loading subscriptions: {str(e)}')
            return redirect('subscription')

    def post(self, request, *args, **kwargs):
        logger.info("Processing subscription deletion POST request")
        try:
            subscription_id = request.POST.get('subscription_id')
            if not subscription_id:
                logger.error("No subscription ID provided")
                return JsonResponse({
                    'status': 'error',
                    'message': 'No subscription selected'
                }, status=400)

            # Get subscription by ID
            subscription = Subscription.objects.get(id=str(subscription_id))
            subscription_name = subscription.name
            subscription.delete()
            logger.info(f"Deleted subscription: {subscription_name}")

            messages.success(request, f'Subscription "{subscription_name}" deleted successfully')
            return JsonResponse({
                'status': 'success', 
                'message': 'Subscription deleted successfully',
                'redirect_url': '/masteradmin/subscription/'
            })

        except Subscription.DoesNotExist:
            logger.error(f"Subscription not found: {subscription_id}")
            return JsonResponse({
                'status': 'error',
                'message': 'Subscription not found'
            }, status=404)

        except Exception as e:
            logger.error(f"Error deleting subscription: {str(e)}")
            return JsonResponse({
                'status': 'error',
                'message': f'Error deleting subscription: {str(e)}'
            }, status=400)

class EditSubscriptionView(View):
    def get(self, request, *args, **kwargs):
        """
        Handles GET requests to display the subscription edit form.
        Retrieves all subscriptions and subscription types for display.
        """
        logger.info("Processing subscription edit GET request")
        try:
            # Get all subscriptions to display in list
            subscriptions = Subscription.objects.all()
            total_subscriptions = subscriptions.count()
            
            # Get subscription types for dropdown
            subscription_types = ['Monthly', 'Quarterly', 'Yearly']
            
            context = {
                'subscriptions': subscriptions,
                'total_subscriptions': total_subscriptions,
                'subscription_types': subscription_types,
                'total_agents': AgentUser.objects.filter(name__isnull=False, is_approved=False).count() + Employee.objects.filter(name__isnull=False, is_approved=False).count() + SupportUser.objects.filter(name__isnull=False, is_approved=False).count(),
                'total_meetings': Meeting.objects.filter(is_active=True).count(),
                'departments': Policy.objects.all(),
                'support_departments': SupportDepartment.objects.all()
            }
            logger.debug(f"Found {total_subscriptions} subscriptions")
            return render(request, 'masteradmin/edit_subscription.html', context)
        except Exception as e:
            logger.error(f"Error loading subscriptions: {str(e)}")
            messages.error(request, f'Error loading subscriptions: {str(e)}')
            return redirect('subscription')

    def post(self, request, *args, **kwargs):
        """
        Handles POST requests to update an existing subscription.
        Validates and updates subscription details including name, price, plan type etc.
        """
        logger.info("Processing subscription edit POST request")
        try:
            # Get all required fields from POST data
            subscription_id = request.POST.get('subscription_id')
            name = request.POST.get('name')
            price_monthly = request.POST.get('price_monthly')
            plan_type = request.POST.get('plan_type')
            validity_days = request.POST.get('validity_days')

            # Validate required fields
            required_fields = {
                'subscription_id': subscription_id,
                'name': name, 
                'price_monthly': price_monthly,
                'plan_type': plan_type,
                'validity_days': validity_days
            }

            missing_fields = [field for field, value in required_fields.items() if not value]
            if missing_fields:
                logger.error(f"Missing required fields: {missing_fields}")
                return JsonResponse({
                    'status': 'error',
                    'message': f'Missing required fields: {", ".join(missing_fields)}'
                }, status=400)

            # Validate numeric fields
            try:
                price_monthly = float(price_monthly)
                validity_days = int(validity_days)
            except ValueError:
                return JsonResponse({
                    'status': 'error',
                    'message': 'Price must be a valid number and validity days must be an integer'
                }, status=400)

            # Validate plan type
            valid_plan_types = ['Monthly', 'Quarterly', 'Yearly']
            if plan_type not in valid_plan_types:
                return JsonResponse({
                    'status': 'error',
                    'message': f'Invalid plan type. Must be one of: {", ".join(valid_plan_types)}'
                }, status=400)

            # Get subscription by ID
            subscription = Subscription.objects.get(id=str(subscription_id))
            
            # Update all fields
            subscription.name = name
            subscription.price_monthly = price_monthly
            subscription.plan_type = plan_type
            subscription.validity_days = validity_days
            subscription.save()

            messages.success(request, f'Subscription "{name}" updated successfully')
            return JsonResponse({
                'status': 'success',
                'message': 'Subscription updated successfully',
                'redirect_url': '/masteradmin/subscription/'
            })

        except Subscription.DoesNotExist:
            return JsonResponse({
                'status': 'error', 
                'message': 'Subscription not found'
            }, status=404)

        except Exception as e:
            logger.error(f"Error updating subscription: {str(e)}")
            return JsonResponse({
                'status': 'error',
                'message': f'Error updating subscription: {str(e)}'
            }, status=400)
class TicketView(TemplateView):
    template_name = "masteradmin/ticket.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['tickets'] = Tickets.objects.all()
        context['total_agents'] = AgentUser.objects.filter(name__isnull=False, is_approved=False).count() + Employee.objects.filter(name__isnull=False, is_approved=False).count() + SupportUser.objects.filter(name__isnull=False, is_approved=False).count()
        context['total_meetings'] = Meeting.objects.filter(is_active=True).count()
        context['departments'] = Policy.objects.all()
        # Get current timestamp and calculate 24 hours ago
        now = timezone.now()
        last_24_hours = now - timedelta(hours=24)

        # Recently opened tickets (last 24 hours)
        context['recent_tickets'] = Tickets.objects.filter(
            created_at__gte=last_24_hours
        ).count()

        # Pending tickets awaiting response
        context['pending_tickets'] = Tickets.objects.filter(
            status='Pending'
        ).count()

        # Priority/urgent tickets
        context['priority_tickets'] = Tickets.objects.filter(
            priority='Urgent'
        ).count()

        # Tickets nearing expiry (within next 24 hours)
        expiry_threshold = now + timedelta(hours=24)
        context['expiring_tickets'] = Tickets.objects.filter(
            created_at__lte=now - timedelta(hours=72),
            status='Pending'
        ).count()
        context['support_departments'] = SupportDepartment.objects.all()
        return context
class CreateCouponView(TemplateView):
    template_name = "masteradmin/create_coupon.html"

    def get_context_data(self, **kwargs):
        logger.info("Fetching all subscriptions for coupon creation form")
        context = super().get_context_data(**kwargs)
        # Get all subscriptions to associate coupons with
        # context['subscriptions'] = Subscription.objects.filter(is_active=True)
        context['coupons'] = Coupons.objects.all()
        context['total_agents'] = AgentUser.objects.filter(is_active=True).count() + Employee.objects.filter(is_active=True).count() + SupportUser.objects.filter(is_active=True).count()
        context['total_meetings'] = Meeting.objects.filter(is_active=True).count()
        context['departments'] = Policy.objects.all()
        context['support_departments'] = SupportDepartment.objects.all()
        logger.debug(f"Found {context['coupons'].count()} coupons")
        return context

    def post(self, request, *args, **kwargs):
        try:
            logger.info(f"Processing coupon creation request from user {request.user}")
            
            code = request.POST.get('code')
            discount_type = request.POST.get('discount_type')
            rate = request.POST.get('value')
            expiry_date = request.POST.get('expiry_date')
            number_of_uses = request.POST.get('max_uses')
            name = request.POST.get('code')  # Using code as name

            logger.debug(f"Received coupon data - Code: {code}, Type: {discount_type}, Rate: {rate}, Expiry: {expiry_date}, Uses: {number_of_uses}")

            # Validate required fields
            if not all([code, discount_type, rate, number_of_uses]):
                logger.warning("Missing required fields in coupon creation request")
                return JsonResponse({
                    'status': 'error',
                    'message': 'All fields are required'
                }, status=400)

            # Validate numeric values
            try:
                rate = float(rate)
                number_of_uses = int(number_of_uses)
                logger.debug(f"Validated numeric values - Rate: {rate}, Uses: {number_of_uses}")
            except ValueError:
                logger.warning(f"Invalid numeric values - Rate: {rate}, Uses: {number_of_uses}")
                return JsonResponse({
                    'status': 'error', 
                    'message': 'Invalid rate or number of uses'
                }, status=400)

            # Create the coupon
            coupon = Coupons.objects.create(
                name=name,
                code=code,
                discount_type=discount_type,
                rate=rate,
                number_of_uses=number_of_uses,
                expiry_date=expiry_date if expiry_date else None
            )
            logger.info(f"Successfully created coupon with ID {coupon.id}")

            messages.success(request, f'Coupon "{code}" created successfully')
            return JsonResponse({
                'status': 'success',
                'message': 'Coupon created successfully',
                'redirect_url': '/masteradmin/subscription/'
            })

        except Exception as e:
            logger.error(f"Error creating coupon: {str(e)}", exc_info=True)
            return JsonResponse({
                'status': 'error',
                'message': f'Error creating coupon: {str(e)}'
            }, status=400)

class ExportDataView(View):
    def get(self, request, *args, **kwargs):
        # Create a new workbook and select the active sheet
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Coupons Data"

        # Define headers
        headers = ['Name', 'Code', 'Discount Type', 'Rate', 'Number of Uses', 'Expiry Date']
        for col, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col, value=header)

        # Get all coupons data
        coupons = Coupons.objects.all()

        # Add data to sheet
        for row, coupon in enumerate(coupons, start=2):
            sheet.cell(row=row, column=1, value=coupon.name)
            sheet.cell(row=row, column=2, value=coupon.code)
            sheet.cell(row=row, column=3, value=coupon.discount_type)
            sheet.cell(row=row, column=4, value=float(coupon.rate))
            sheet.cell(row=row, column=5, value=coupon.number_of_uses)
            sheet.cell(row=row, column=6, value=str(coupon.expiry_date) if coupon.expiry_date else "No expiry")

        # Create response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=coupons_data.xlsx'

        # Save workbook to response
        workbook.save(response)
        return response

@method_decorator(csrf_exempt, name='dispatch')
class DeleteCouponView(View):
    def post(self, request):
        try:
            # Check if request body is empty
            if not request.body:
                return JsonResponse({
                    'status': 'error',
                    'message': 'Request body is required'
                }, status=400)

            try:
                data = json.loads(request.body)
            except json.JSONDecodeError:
                return JsonResponse({
                    'status': 'error',
                    'message': 'Invalid JSON in request body'
                }, status=400)

            coupon_id = data.get('coupon_id')
            
            if not coupon_id:
                return JsonResponse({
                    'status': 'error',
                    'message': 'Coupon ID is required'
                }, status=400)

            coupon = Coupons.objects.get(id=coupon_id)
            coupon_code = coupon.code
            coupon.delete()

            messages.success(request, f'Coupon "{coupon_code}" deleted successfully')
            return JsonResponse({
                'status': 'success',
                'message': 'Coupon deleted successfully',
                'redirect_url': '/masteradmin/subscription/'
            })

        except Coupons.DoesNotExist:
            return JsonResponse({
                'status': 'error', 
                'message': 'Coupon not found'
            }, status=404)
            
        except Exception as e:
            logger.error(f"Error deleting coupon: {str(e)}", exc_info=True)
            return JsonResponse({
                'status': 'error',
                'message': f'Error deleting coupon: {str(e)}'
            }, status=400)
        

class AgentDetailsView(TemplateView):
    template_name = "masteradmin/agent_details.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['agents'] = AgentUser.objects.exclude(password__isnull=True).exclude(password='')
        context['supports'] = SupportUser.objects.exclude(password__isnull=True).exclude(password='')
        context['employees'] = Employee.objects.exclude(password__isnull=True).exclude(password='')
        context['total_agents'] = AgentUser.objects.filter(name__isnull=False, is_approved=False).count() + Employee.objects.filter(name__isnull=False, is_approved=False).count() + SupportUser.objects.filter(name__isnull=False, is_approved=False).count()
        context['total_meetings'] = Meeting.objects.filter(is_active=True).count()
        context['departments'] = Policy.objects.all()
        context['support_departments'] = SupportDepartment.objects.all()
        return context
    
class ApproveAgentView(View):
    def dispatch(self, request, *args, **kwargs):
        if request.method != 'POST':
            return JsonResponse({"status": "error", "message": "Method Not Allowed"}, status=405)
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        agent_id = self.kwargs.get('agent_id')
        try:
            agent = AgentUser.objects.get(id=agent_id)
            if agent.is_approved:
                return JsonResponse({'status': 'error', 'message': 'Agent is already approved'}, status=400)
            
            agent.is_approved = True
            agent.is_active = True
            agent.save()

            # Send approval email to agent
            try:
                subject = 'Your Account Has Been Approved'
                message = f"""
                Dear {agent.name},

                Your account has been approved. You can now login to your account using your registered email and password.

                Thank you for joining us!

                login : https://{request.get_host()}/agents/

                Best regards,
                The Admin Team
                """
                send_mail(
                    subject,
                    message,
                    settings.DEFAULT_FROM_EMAIL,
                    [agent.email],
                    fail_silently=False,
                )
                logger.info(f"Approval email sent successfully to agent {agent.email}")
            except Exception as e:
                logger.error(f"Failed to send approval email to {agent.email}: {str(e)}")
                # Continue execution even if email fails
                
            return JsonResponse({'status': 'success', 'message': 'Agent approved successfully'})
        except AgentUser.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Agent not found'}, status=404)

class DepartmentsView(TemplateView):
    template_name = "masteradmin/departments.html"

    def get_context_data(self, **kwargs):
        print("Getting context data for departments view")
        context = super().get_context_data(**kwargs)
        departments = Policy.objects.all()
        context['total_agents'] = AgentUser.objects.filter(name__isnull=False, is_approved=False).count() + Employee.objects.filter(name__isnull=False, is_approved=False).count() + SupportUser.objects.filter(name__isnull=False, is_approved=False).count()
        context['total_meetings'] = Meeting.objects.filter(is_active=True).count()
        context['support_departments'] = SupportDepartment.objects.all()
        print(f"Found {departments.count()} departments")
        context['departments'] = departments
        return context

    def post(self, request, *args, **kwargs):
        print("Handling POST request to create department")
        try:
            data = json.loads(request.body)
            print(f"Received data: {data}")
            department_name = data.get('department_name')
            department_terms = data.get('department_terms')

            if not department_name:
                print("Department name is missing")
                return JsonResponse({
                    'status': 'error',
                    'message': 'Department name is required'
                }, status=400)

            if not department_terms:
                print("Department terms are missing")
                return JsonResponse({
                    'status': 'error', 
                    'message': 'Department terms are required'
                }, status=400)

            print(f"Creating department with name: {department_name}")
            department = Policy.objects.create(
                name=department_name,
                terms_and_conditions=department_terms
            )
            print(f"Successfully created department with ID: {department.id}")

            return JsonResponse({
                'status': 'success',
                'message': 'Department created successfully',
                'department': {
                    'id': department.id,
                    'name': department.name,
                    'terms_and_conditions': department.terms_and_conditions
                }
            })

        except json.JSONDecodeError:
            print("Invalid JSON data received")
            return JsonResponse({
                'status': 'error',
                'message': 'Invalid JSON data'
            }, status=400)

        except Exception as e:
            print(f"Error creating department: {str(e)}")
            logger.error(f"Error creating department: {str(e)}", exc_info=True)
            return JsonResponse({
                'status': 'error',
                'message': 'Error creating department'
            }, status=500)

    def delete(self, request, *args, **kwargs):
        print("Handling DELETE request for department")
        try:
            department_id = kwargs.get('pk')
            print(f"Looking for department ID: {department_id}")
            department = Policy.objects.get(id=department_id)
            print(f"Found department: {department.name}")
            department.delete()
            print("Department deleted successfully")
            return JsonResponse({
                'status': 'success',
                'message': 'Department deleted successfully'
            })
        except Policy.DoesNotExist:
            print(f"Department with ID {department_id} not found")
            return JsonResponse({
                'status': 'error',
                'message': 'Department not found'
            }, status=404)
        except Exception as e:
            print(f"Error deleting department: {str(e)}")
            logger.error(f"Error deleting department: {str(e)}", exc_info=True)
            return JsonResponse({
                'status': 'error', 
                'message': 'Error deleting department'
            }, status=500)

    def put(self, request, department_id, *args, **kwargs):
        print(f"Handling PUT request for department ID: {department_id}")
        try:
            data = json.loads(request.body)
            print(f"Received data: {data}")
            department = Policy.objects.get(id=department_id)
            print(f"Found department: {department.name}")

            department_name = data.get('department_name')
            department_terms = data.get('department_terms')

            if not department_name:
                print("Department name is missing")
                return JsonResponse({
                    'status': 'error',
                    'message': 'Department name is required'
                }, status=400)

            if not department_terms:
                print("Department terms are missing")
                return JsonResponse({
                    'status': 'error',
                    'message': 'Department terms are required'
                }, status=400)

            department.name = department_name
            department.terms_and_conditions = department_terms
            department.save()
            print(f"Successfully updated department: {department.name}")

            return JsonResponse({
                'status': 'success',
                'message': 'Department updated successfully',
                'department': {
                    'id': department.id,
                    'name': department.name,
                    'terms_and_conditions': department.terms_and_conditions
                }
            })

        except Policy.DoesNotExist:
            print(f"Department with ID {department_id} not found")
            return JsonResponse({
                'status': 'error',
                'message': 'Department not found'
            }, status=404)
        except json.JSONDecodeError:
            print("Invalid JSON data received")
            return JsonResponse({
                'status': 'error',
                'message': 'Invalid JSON data'
            }, status=400)
        except Exception as e:
            print(f"Error updating department: {str(e)}")
            logger.error(f"Error updating department: {str(e)}", exc_info=True)
            return JsonResponse({
                'status': 'error',
                'message': 'Error updating department'
            }, status=500)
        
class CustomerSupportMAView(TemplateView):
    template_name = "masteradmin/customer_support2.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['total_agents'] = AgentUser.objects.filter(name__isnull=False, is_approved=False).count() + Employee.objects.filter(name__isnull=False, is_approved=False).count() + SupportUser.objects.filter(name__isnull=False, is_approved=False).count()
        context['total_meetings'] = Meeting.objects.filter(is_active=True).count()
        context['departments'] = Policy.objects.all()
        context['support_departments'] = SupportDepartment.objects.all()
        return context

class AllEmployeesView(TemplateView):
    template_name = "masteradmin/allemployees.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['employees'] = Employee.objects.filter(is_active=True)
        context['departments'] = Policy.objects.all()
        context['total_agents'] = AgentUser.objects.filter(name__isnull=False, is_approved=False).count() + Employee.objects.filter(name__isnull=False, is_approved=False).count() + SupportUser.objects.filter(name__isnull=False, is_approved=False).count()
        context['total_meetings'] = Meeting.objects.filter(is_active=True).count()
        context['support_departments'] = SupportDepartment.objects.all()
        return context

class CreateEmployeeView(View):
    def post(self, request):
        print("Handling POST request to create employee")
        try:
            data = json.loads(request.body)
            email = data.get('email')
            department_id = data.get('department')
            print(f"Received data - email: {email}, department_id: {department_id}")

            # Basic validation
            if not email or not department_id:
                return JsonResponse({
                    'status': 'error',
                    'message': 'Email and department are required fields'
                }, status=400)

            # Validate email format
            if not re.match(r"[^@]+@[^@]+\.[^@]+", email):
                return JsonResponse({
                    'status': 'error',
                    'message': 'Invalid email format'
                }, status=400)

            # Check if employee exists
            if Employee.objects.filter(email=email).exists():
                return JsonResponse({
                    'status': 'error',
                    'message': 'An employee with this email already exists'
                }, status=400)

            # Get department
            try:
                department = Policy.objects.get(id=department_id)
            except Policy.DoesNotExist:
                return JsonResponse({
                    'status': 'error', 
                    'message': 'Invalid department selected'
                }, status=404)
            except ValueError:
                return JsonResponse({
                    'status': 'error',
                    'message': 'Invalid department ID format'
                }, status=400)

            # Generate unique passcode
            while True:
                passcode = ''.join(random.choices('0123456789', k=4))
                if not Employee.objects.filter(employee_passcode=passcode).exists():
                    break

            # Create employee
            try:
                employee = Employee.objects.create(
                    email=email,
                    employee_passcode=passcode,
                    department=department,
                    is_active=False
                )
            except IntegrityError:
                return JsonResponse({
                    'status': 'error',
                    'message': 'Error creating employee record'
                }, status=500)

            # Generate registration link
            registration_link = f"{request.get_host()}/employee/register/{passcode}/"

            # Send email
            try:
                email_subject = 'Complete Your Employee Registration'
                email_body = f"""
                Hello,

                You have been invited to join as an employee. Please use the following details to complete your registration:

                Passcode: {passcode}

                Click here to complete your registration:
                {registration_link}

                This link will expire in 24 hours.

                Best regards,
                Your Organization
                """

                send_mail(
                    email_subject,
                    email_body,
                    settings.DEFAULT_FROM_EMAIL,
                    [email],
                    fail_silently=False,
                )
            except Exception as e:
                # Rollback employee creation if email fails
                employee.delete()
                logger.error(f"Failed to send registration email: {str(e)}", exc_info=True)
                return JsonResponse({
                    'status': 'error',
                    'message': 'Failed to send registration email'
                }, status=500)

            return JsonResponse({
                'status': 'success',
                'message': 'Registration invitation sent successfully',
                'data': {
                    'employee_id': str(employee.id),
                    'email': employee.email,
                    'department': department.name,
                    'passcode': passcode
                }
            })

        except json.JSONDecodeError:
            return JsonResponse({
                'status': 'error',
                'message': 'Invalid request format'
            }, status=400)
        except Exception as e:
            logger.error(f"Unexpected error creating employee: {str(e)}", exc_info=True)
            return JsonResponse({
                'status': 'error',
                'message': 'An unexpected error occurred'
            }, status=500)

class MasterSupportView(TemplateView):
    template_name = "masteradmin/mastersupport.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['support_departments'] = SupportDepartment.objects.all()
        context['total_agents'] = AgentUser.objects.filter(name__isnull=False, is_approved=False).count() + Employee.objects.filter(name__isnull=False, is_approved=False).count() + SupportUser.objects.filter(name__isnull=False, is_approved=False).count()
        context['total_meetings'] = Meeting.objects.filter(is_active=True).count()
        context['policies'] = Policy.objects.all()
        context['support_users'] = SupportUser.objects.all()
        return context

@method_decorator(csrf_exempt, name='dispatch')
class AddSupportView(View):
    def post(self, request):
        logger.info("AddSupportView.post() called")
        try:
            if not request.body:
                logger.error("Empty request body received")
                return JsonResponse({"status": "error", "message": "Empty request body"}, status=400)

            # Parse JSON data
            try:
                data = json.loads(request.body.decode('utf-8'))
                logger.info(f"Request body: {request.body}")
                logger.info(f"Parsed data: {data}")
            except json.JSONDecodeError as e:
                logger.error(f"JSON decode error: {str(e)}")
                # Try to get form data instead
                data = request.POST
                if not data:
                    return JsonResponse({"status": "error", "message": "Invalid JSON format and no form data found"}, status=400)
                logger.info(f"Using form data instead: {data}")
            
            support_email = data.get('email')
            department = data.get('department')
            policy = data.get('policy')

            if not all([support_email, department, policy]):
                logger.error(f"Missing required fields - email: {bool(support_email)}, department: {bool(department)}, policy: {bool(policy)}")
                return JsonResponse({"status": "error", "message": "Missing required fields"}, status=400)

            try:
                department_instance = SupportDepartment.objects.get(id=department)
                policy_instance = Policy.objects.get(id=policy)
            except SupportDepartment.DoesNotExist:
                logger.error(f"Support department with id {department} does not exist")
                return JsonResponse({"status": "error", "message": "Support department not found"}, status=400)
            except Policy.DoesNotExist:
                logger.error(f"Policy with id {policy} does not exist")
                return JsonResponse({"status": "error", "message": "Policy not found"}, status=400)

            print(policy_instance)
            logger.info(f"Extracted fields - email: {support_email}, department: {department}")

            # Check if email already exists
            if SupportUser.objects.filter(email=support_email).exists():
                logger.error(f"Support with email {support_email} already exists")
                return JsonResponse({"status": "error", "message": "Email already registered"}, status=400)

            # Generate new passcode
            last_support = SupportUser.objects.order_by('-support_user_passcode').first()
            if last_support and last_support.support_user_passcode:
                new_passcode = int(last_support.support_user_passcode) + 1
            else:
                new_passcode = 7000

            # Create new agent with only required fields, rest as null
            try:
                support = SupportUser.objects.create(
                    email=support_email,
                    support_user_passcode=str(new_passcode),
                    support_department=department_instance,
                    terms_and_conditions=policy_instance,
                    photo=None,
                    name=None,
                    password=None,
                    phone_number=None,
                    pan_number=None,
                    aadhar_number=None,
                    address=None,
                    dob=None,
                    qualification=None,
                    qualification_file=None,
                    bank_account_holder_name=None,
                    bank_account_number=None,
                    bank_name=None,
                    branch_name=None,
                    bank_ifsc_code=None,
                    cancelled_cheque_file=None,
                    offer_letter_file=None,
                    bank_statement_file=None,
                    increment_letter_file=None,
                    pay_slip_file=None,
                    experience_letter_file=None,
                    leave_letter_file=None,
                    addhar_card_file=None
                )

                # Generate form link with passcode
                form_link = f"http://{request.get_host()}/customersupport/register/{new_passcode}/"
                
                # Send email with form link
                email_subject = "Complete Your Support Registration"
                email_message = f"""
                Hello,

                Please complete your support registration by clicking the link below:
                {form_link}

                Your passcode: {new_passcode}

                Best regards,
                1Matrix Team
                """

                # Add proper email sending with error handling and retries
                max_retries = 3
                retry_count = 0
                email_sent = False

                while retry_count < max_retries and not email_sent:
                    try:
                        send_mail(
                            subject=email_subject,
                            message=email_message,
                            from_email=settings.EMAIL_HOST_USER,
                            recipient_list=[support_email],
                            fail_silently=False,
                            auth_user=settings.EMAIL_HOST_USER,
                            auth_password=settings.EMAIL_HOST_PASSWORD
                        )
                        email_sent = True
                        logger.info(f"Registration email sent successfully to {support_email}")
                    except Exception as e:
                        retry_count += 1
                        logger.error(f"Email sending attempt {retry_count} failed: {str(e)}")
                        if retry_count == max_retries:
                            # Delete the agent if email sending completely fails
                            support.delete()
                            logger.error("Max retries reached for email sending. Support creation rolled back.")
                            return JsonResponse({
                                "status": "error", 
                                "message": "Failed to send registration email. Please try again."
                            }, status=500)

                logger.info(f"Support created successfully with email: {support_email} and passcode: {new_passcode}")
                return JsonResponse({
                    "status": "success", 
                    "message": "Support added successfully! Registration link has been sent to the email."
                })

            except Exception as e:
                logger.error(f"Error creating support: {str(e)}")
                return JsonResponse({"status": "error", "message": "Error creating support"}, status=500)

        except Exception as e:
            logger.error(f"Unexpected error: {str(e)}")
            return JsonResponse({"status": "error", "message": str(e)}, status=500)

        
class CreateNoteView(TemplateView):
    template_name = "masteradmin/create_note.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['total_agents'] = AgentUser.objects.filter(name__isnull=False, is_approved=False).count() + Employee.objects.filter(name__isnull=False, is_approved=False).count() + SupportUser.objects.filter(name__isnull=False, is_approved=False).count()
        context['total_meetings'] = Meeting.objects.filter(is_active=True).count()
        context['departments'] = Policy.objects.all()
        context['support_departments'] = SupportDepartment.objects.all()
        context['quick_notes'] = QuickNotes.objects.all().order_by('-is_pinned', '-created_at')
        return context

    @method_decorator(csrf_exempt)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def post(self, request, *args, **kwargs):
        try:
            # Print request data for debugging
            print("Request Content Type:", request.content_type)
            print("Request Body:", request.body)
            print("Request POST:", request.POST)

            # Handle both JSON and form data
            if request.content_type and 'application/json' in request.content_type:
                try:
                    data = json.loads(request.body)
                    title = data.get('title')
                    content = data.get('note')  # Changed from 'content' to 'note' to match model
                except json.JSONDecodeError as e:
                    print("JSON Decode Error:", str(e))
                    return JsonResponse({
                        'status': 'error',
                        'message': 'Invalid JSON data'
                    }, status=400)
            else:
                title = request.POST.get('title')
                content = request.POST.get('note')  # Changed from 'content' to 'note'

            print("Title:", title)
            print("Content:", content)

            # Validate input
            if not title or not content:
                error_message = 'Title and content are required'
                print("Validation Error:", error_message)
                return JsonResponse({
                    'status': 'error',
                    'message': error_message
                }, status=400)

            # Create and save the note
            quick_note = QuickNotes(
                title=title,
                note=content,
                is_pinned=False,
                created_at=timezone.now()
            )
            quick_note.save()

            print("Note saved successfully:", quick_note.id)

            # Return appropriate response
            if request.content_type and 'application/json' in request.content_type:
                return JsonResponse({
                    'status': 'success',
                    'message': 'Note created successfully',
                    'note': {
                        'id': str(quick_note.id),
                        'title': quick_note.title,
                        'note': quick_note.note,
                        'is_pinned': quick_note.is_pinned,
                        'created_at': quick_note.created_at.isoformat()
                    }
                })
            else:
                messages.success(request, 'Note created successfully')
                return redirect('create_note')

        except Exception as e:
            print("Error creating note:", str(e))
            logger.error(f"Error creating note: {str(e)}", exc_info=True)
            error_message = f'Error creating note: {str(e)}'
            if request.content_type and 'application/json' in request.content_type:
                return JsonResponse({
                    'status': 'error',
                    'message': error_message
                }, status=500)
            else:
                messages.error(request, error_message)
                return redirect('create_note')

    def delete(self, request, *args, **kwargs):
        try:
            note_id = kwargs.get('note_id')
            note = QuickNotes.objects.get(id=note_id)
            note.delete()
            return JsonResponse({'status': 'success'})
        except QuickNotes.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Note not found'}, status=404)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
        
class MarkQuickNoteAsPinnedView(View):
    def post(self, request, *args, **kwargs):
        note_id = self.kwargs.get('note_id')
        try:
            note = QuickNotes.objects.get(id=note_id)
            note.is_pinned = not note.is_pinned
            note.save()
            return JsonResponse({'status': 'success', 'is_pinned': note.is_pinned})
        except QuickNotes.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Note not found'}, status=404)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

class RejectAgentView(View):
    def dispatch(self, request, *args, **kwargs):
        if request.method != 'POST':
            return JsonResponse({"status": "error", "message": "Method Not Allowed"}, status=405)
        return super().dispatch(request, *args, **kwargs)
    def post(self, request, *args, **kwargs):
        agent_id = self.kwargs.get('agent_id')
        try:
            agent = AgentUser.objects.get(id=agent_id)
            if agent.is_rejected:
                return JsonResponse({'status': 'error', 'message': 'Agent is already rejected'}, status=400)
            
            agent.is_approved = False
            agent.is_active = False
            agent.is_rejected = True
            agent.save()

            # Send rejection email to agent
            try:
                subject = 'Your Account Has Been Rejected'
                message = f"""
                Dear {agent.name},

                We regret to inform you that your account registration has been rejected.

                If you believe this was done in error, please contact our support team.

                Best regards,
                The Admin Team
                1Matrix
                """
                send_mail(
                    subject,
                    message,
                    settings.DEFAULT_FROM_EMAIL,
                    [agent.email],
                    fail_silently=False,
                )
                logger.info(f"Rejection email sent successfully to agent {agent.email}")
            except Exception as e:
                logger.error(f"Failed to send rejection email to {agent.email}: {str(e)}")
                # Continue execution even if email fails
                
            return JsonResponse({'status': 'success', 'message': 'Agent rejected successfully'})
        except AgentUser.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Agent not found'}, status=404)

class MarkEmployeeAsActiveView(View):
    def get(self, request, employee_id):
        try:
            employee = Employee.objects.get(id=employee_id)
            if employee.is_active:
                return JsonResponse({'status': 'info', 'message': 'Employee is already active', 'redirect': request.META.get('HTTP_REFERER')})
            else:
                employee.is_active = True
                employee.save()
                return JsonResponse({'status': 'success', 'message': 'Employee can be marked as active', 'redirect': request.META.get('HTTP_REFERER')})
        except Employee.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Employee not found'}, status=404)
    def post(self, request, *args, **kwargs):
        employee_id = self.kwargs.get('employee_id')
        try:
            employee = Employee.objects.get(id=employee_id)
            if employee.is_approved:
                return JsonResponse({'status': 'error', 'message': 'Employee is already approved'}, status=400)
            
            employee.is_approved = True
            employee.is_active = True
            employee.save()

            # Send approval email to employee
            try:
                subject = 'Your Account Has Been Approved'
                message = f"""
                Dear {employee.name},

                Your account has been approved. You can now login to your account using your registered email and password.

                Thank you for joining us!

                login : https://{request.get_host()}/employee/

                Best regards,
                The Admin Team
                1Matrix
                """
                send_mail(
                    subject,
                    message,
                    settings.DEFAULT_FROM_EMAIL,
                    [employee.email],
                    fail_silently=False,
                )
                logger.info(f"Approval email sent successfully to employee {employee.email}")
            except Exception as e:
                logger.error(f"Failed to send approval email to {employee.email}: {str(e)}")
                # Continue execution even if email fails
                
            return JsonResponse({'status': 'success', 'message': 'Employee approved successfully'})
        except Employee.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Employee not found'}, status=404)
        

class MarkEmployeeAsInactiveView(View):
    def post(self, request, *args, **kwargs):
        employee_id = self.kwargs.get('employee_id')
        try:
            employee = Employee.objects.get(id=employee_id)
            if not employee.is_active:
                return JsonResponse({'status': 'error', 'message': 'Employee is already inactive'}, status=400)
            
            employee.is_active = False
            employee.save()

            # Send deactivation email to employee
            try:
                subject = 'Your Account Has Been Deactivated'
                message = f"""
                Dear {employee.name},

                Your account has been deactivated. You will no longer be able to login until your account is reactivated.

                If you have any questions, please contact the admin team.

                Best regards,
                The Admin Team
                1Matrix
                """
                send_mail(
                    subject,
                    message,
                    settings.DEFAULT_FROM_EMAIL,
                    [employee.email],
                    fail_silently=False,
                )
                logger.info(f"Deactivation email sent successfully to employee {employee.email}")
            except Exception as e:
                logger.error(f"Failed to send deactivation email to {employee.email}: {str(e)}")
                # Continue execution even if email fails
                
            return JsonResponse({'status': 'success', 'message': 'Employee deactivated successfully'})
        except Employee.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Employee not found'}, status=404)
        

class RejectEmployeeView(View):
    def post(self, request, *args, **kwargs):
        employee_id = self.kwargs.get('employee_id')
        try:
            employee = Employee.objects.get(id=employee_id)
            if not employee.is_approved:
                return JsonResponse({'status': 'error', 'message': 'Employee is already rejected'}, status=400)
            
            employee.is_approved = False
            employee.is_active = False
            employee.save()

            # Send rejection email to employee
            try:
                subject = 'Your Account Has Been Rejected'
                message = f"""
                Dear {employee.name},

                We regret to inform you that your account registration has been rejected.

                If you have any questions, please contact the admin team.

                Best regards,
                The Admin Team
                1Matrix
                """
                send_mail(
                    subject,
                    message,
                    settings.DEFAULT_FROM_EMAIL,
                    [employee.email],
                    fail_silently=False,
                )
                logger.info(f"Rejection email sent successfully to employee {employee.email}")
            except Exception as e:
                logger.error(f"Failed to send rejection email to {employee.email}: {str(e)}")
                # Continue execution even if email fails
                
            return JsonResponse({'status': 'success', 'message': 'Employee rejected successfully'})
        except Employee.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Employee not found'}, status=404)

class DeleteDepartmentView(View):
    def delete(self, request, department_id):
        try:
            department = Policy.objects.get(id=department_id)
            department.delete()
            return JsonResponse({'status': 'success', 'message': 'Department deleted successfully'})
        except Policy.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Department not found'}, status=404)
        except Exception as e:
            logger.error(f"Error deleting department: {str(e)}", exc_info=True)
            return JsonResponse({'status': 'error', 'message': 'An error occurred while deleting the department'}, status=500)

class SuspendedAgentsView(TemplateView):
    template_name = "masteradmin/suspended_agents.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['agents'] = AgentUser.objects.filter(is_suspended=True)
        context['total_agents'] = AgentUser.objects.filter(name__isnull=False, is_approved=False).count() + Employee.objects.filter(name__isnull=False, is_approved=False).count() + SupportUser.objects.filter(name__isnull=False, is_approved=False).count()
        context['total_meetings'] = Meeting.objects.filter(is_active=True).count()
        context['departments'] = Policy.objects.all()
        context['support_departments'] = SupportDepartment.objects.all()
        return context

class AllAgentsView(TemplateView):
    template_name = "masteradmin/all_agents.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['agents'] = AgentUser.objects.all()
        context['total_agents'] = AgentUser.objects.filter(name__isnull=False, is_approved=False).count() + Employee.objects.filter(name__isnull=False, is_approved=False).count() + SupportUser.objects.filter(name__isnull=False, is_approved=False).count()
        context['total_meetings'] = Meeting.objects.filter(is_active=True).count()
        context['departments'] = Policy.objects.all()
        context['support_departments'] = SupportDepartment.objects.all()
        context['policies'] = Policy.objects.all()
        return context

class AgentPerformanceView(TemplateView):
    template_name = "masteradmin/agent_performance.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['agents'] = AgentUser.objects.all()
        context['total_agents'] = AgentUser.objects.filter(name__isnull=False, is_approved=False).count() + Employee.objects.filter(name__isnull=False, is_approved=False).count() + SupportUser.objects.filter(name__isnull=False, is_approved=False).count()
        context['total_meetings'] = Meeting.objects.filter(is_active=True).count()
        context['departments'] = Policy.objects.all()
        context['support_departments'] = SupportDepartment.objects.all()
        return context

class CreateAgentsNotificationView(View):
    def post(self, request):
        try:
            data = json.loads(request.body)
            title = data.get('title')
            message = data.get('message')
            department = data.get('department')
            receivers = data.get('receivers')

            notification = AgentNotification.objects.create(
                title=title,
                message=message,
                department=department,
                receivers=receivers,
                created_at=timezone.now(),
                # updated_at=timezone.now()
            )
            return JsonResponse({'status': 'success', 'message': 'Notification created successfully'})
        except Exception as e:
            logger.error(f"Error creating notification: {str(e)}", exc_info=True)
            return JsonResponse({'status': 'error', 'message': 'An error occurred while creating the notification'}, status=500)

class SendNotificationView(View):
    def post(self, request):
        try:
            data = json.loads(request.body)
            agent_id = data.get('agent_id')
            message = data.get('message')
            agent = AgentUser.objects.get(id=agent_id)
            agent.send_notification(message)
            return JsonResponse({'status': 'success', 'message': 'Notification sent successfully'})
        except Exception as e:
            logger.error(f"Error sending notification: {str(e)}", exc_info=True)
            return JsonResponse({'status': 'error', 'message': 'An error occurred while sending the notification'}, status=500)

class AgentsNotificationView(TemplateView):
    template_name = "masteradmin/agents_notification.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['notifications'] = AgentNotification.objects.all()
        context['total_agents'] = AgentUser.objects.filter(name__isnull=False, is_approved=False).count() + Employee.objects.filter(name__isnull=False, is_approved=False).count() + SupportUser.objects.filter(name__isnull=False, is_approved=False).count()
        context['total_meetings'] = Meeting.objects.filter(is_active=True).count()
        context['departments'] = Policy.objects.all()
        context['support_departments'] = SupportDepartment.objects.all()
        return context
    

class SupportDepartmentView(TemplateView):
    template_name = "masteradmin/support_department.html"

    def get_context_data(self, **kwargs):
        logger.info("Getting context data for SupportDepartmentView")
        context = super().get_context_data(**kwargs)
        departments = SupportDepartment.objects.all()
        logger.debug(f"Found {departments.count()} departments")
        context['total_agents'] = AgentUser.objects.filter(name__isnull=False, is_approved=False).count() + Employee.objects.filter(name__isnull=False, is_approved=False).count() + SupportUser.objects.filter(name__isnull=False, is_approved=False).count()
        context['total_meetings'] = Meeting.objects.filter(is_active=True).count()
        context['policies'] = Policy.objects.all()
        context['support_departments'] = SupportDepartment.objects.all()
        return context
    
    def post(self, request, *args, **kwargs):
        logger.info("Received POST request for support department")
        try:
            action = request.POST.get('action')
            department_id = request.POST.get('department_id')

            if action == 'create':
                name = request.POST.get('name')
                if not name:
                    logger.error("Department name is missing")
                    return JsonResponse({'status': 'error', 'message': 'Department name is required'}, status=400)
                
                department = SupportDepartment.objects.create(name=name)
                logger.info(f"Created department with id: {department.id}")
                return JsonResponse({'status': 'success', 'message': 'Department created successfully'})

            elif action == 'edit':
                name = request.POST.get('name')
                if not name or not department_id:
                    logger.error("Missing required fields for edit")
                    return JsonResponse({'status': 'error', 'message': 'Department name and ID are required'}, status=400)
                
                department = SupportDepartment.objects.get(id=department_id)
                department.name = name
                department.save()
                logger.info(f"Updated department {department_id} with new name: {name}")
                return JsonResponse({'status': 'success', 'message': 'Department updated successfully'})

            elif action == 'delete':
                if not department_id:
                    logger.error("Missing department ID for deletion")
                    return JsonResponse({'status': 'error', 'message': 'Department ID is required'}, status=400)
                
                department = SupportDepartment.objects.get(id=department_id)
                department.delete()
                logger.info(f"Deleted department {department_id}")
                return JsonResponse({'status': 'success', 'message': 'Department deleted successfully'})

            else:
                logger.error(f"Invalid action received: {action}")
                return JsonResponse({'status': 'error', 'message': 'Invalid action'}, status=400)

        except SupportDepartment.DoesNotExist:
            logger.error(f"Department {department_id} not found")
            return JsonResponse({'status': 'error', 'message': 'Department not found'}, status=404)
            
        except Exception as e:
            logger.error(f"Error processing department request: {str(e)}")
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
        
    def put(self, request, *args, **kwargs):
        logger.info("Received PUT request for support department")
        try:
            department_id = request.PUT.get('department_id')
            name = request.PUT.get('department_name')

            if not department_id or not name:
                logger.error("Missing required fields for update")
                return JsonResponse({'status': 'error', 'message': 'Department ID and name are required'}, status=400)

            try:
                department = SupportDepartment.objects.get(id=department_id)
                department.name = name
                department.save()
                logger.info(f"Updated department {department_id} with new name: {name}")
                return JsonResponse({'status': 'success', 'message': 'Department updated successfully'})

            except SupportDepartment.DoesNotExist:
                logger.error(f"Department {department_id} not found")
                return JsonResponse({'status': 'error', 'message': 'Department not found'}, status=404)

        except Exception as e:
            logger.error(f"Error updating department: {str(e)}")
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
            

class WhatsOnMindView(View):
    def post(self, request):
        logger.info("Handling POST request in WhatsOnMindView")
        try:
            data = json.loads(request.body)
            logger.info(f"Received data: {data}")
            
            department_type = data.get('department_type')
            title = data.get('title')
            message = data.get('message')
            logger.info(f"Department type: {department_type}, Title: {title}, Message: {message}")

            if not department_type or not message or not title:
                logger.error("Missing required fields")
                return JsonResponse({
                    'status': 'error',
                    'message': 'Department type, title and message are required'
                }, status=400)

            # Get department name for saving
            department_name = None
            if department_type == 'agents':
                department_name = 'agents'
            elif department_type == 'employees':
                department_name = 'employees'
            elif department_type.startswith('support_'):
                try:
                    department_id = department_type.split('_')[1]
                    department = SupportDepartment.objects.get(id=department_id)
                    department_name = department.name
                except SupportDepartment.DoesNotExist:
                    logger.error(f"Support department with id {department_id} not found")
                    return JsonResponse({
                        'status': 'error',
                        'message': 'Invalid support department'
                    }, status=400)

            # Create main WhatsOnMind record with department name
            logger.info("Creating main WhatsOnMind record")
            whats_on_mind = WhatsOnMind.objects.create(
                title=title,
                message=message,
                department_type=department_name,
                created_at=timezone.now()
            )
            logger.info(f"Created main WhatsOnMind record with id: {whats_on_mind.id}")

            notification_sent = False
            recipients = []

            # Get recipients based on department type
            if department_type == 'agents':
                recipients = AgentUser.objects.filter(is_active=True, is_approved=True)
                logger.info(f"Found {recipients.count()} active agents")
            elif department_type == 'employees':
                recipients = Employee.objects.filter(is_active=True)
                logger.info(f"Found {recipients.count()} active employees")
            elif department_type.startswith('support_'):
                try:
                    department = SupportDepartment.objects.get(id=department_id)
                    recipients = SupportUser.objects.filter(
                        support_department=department,
                        is_active=True,
                        is_approved=True
                    )
                    logger.info(f"Found {recipients.count()} active support users")
                except SupportDepartment.DoesNotExist:
                    logger.error(f"Support department with id {department_id} not found")
                    whats_on_mind.delete()
                    return JsonResponse({
                        'status': 'error',
                        'message': 'Invalid support department'
                    }, status=400)

            # Send notifications to all recipients
            for recipient in recipients:
                try:
                    # Set department_type based on recipient type
                    if isinstance(recipient, AgentUser):
                        whats_on_mind.department_type = 'agents'
                    elif isinstance(recipient, Employee):
                        whats_on_mind.department_type = 'employees'
                    elif isinstance(recipient, SupportUser):
                        whats_on_mind.department_type = recipient.support_department.name
                    
                    whats_on_mind.save()
                    logger.info(f"Processed {recipient.__class__.__name__} {recipient.id}")
                    notification_sent = True
                except Exception as e:
                    logger.error(f"Failed to process {recipient.__class__.__name__} {recipient.id}: {str(e)}")

            if not notification_sent:
                whats_on_mind.delete()
                return JsonResponse({
                    'status': 'error',
                    'message': 'Failed to send notifications to any recipients'
                }, status=500)

            logger.info("Successfully sent all messages")
            return JsonResponse({
                'status': 'success',
                'message': 'Message sent successfully',
                'whats_on_mind': {
                    'id': whats_on_mind.id,
                    'title': whats_on_mind.title,
                    'message': whats_on_mind.message,
                    'department_type': whats_on_mind.department_type,
                    'created_at': whats_on_mind.created_at.isoformat()
                }
            })

        except json.JSONDecodeError:
            logger.error("Invalid JSON data received")
            return JsonResponse({
                'status': 'error', 
                'message': 'Invalid JSON data'
            }, status=400)
        except Exception as e:
            logger.error(f"Error occurred: {str(e)}", exc_info=True)
            return JsonResponse({
                'status': 'error',
                'message': 'An error occurred while sending the message'
            }, status=500)
        

class SendAgentNotificationView(View):
    def post(self, request):
        try:
            print("Received request body:", request.body)
            data = json.loads(request.body)
            agent_ids = data.get('agent_ids', [])
            message = data.get('message')
            print("Parsed data:", data)
            print("Agent IDs:", agent_ids)
            print("Message:", message)

            if not agent_ids or not message:
                print("Missing required fields - agents:", bool(agent_ids), "message:", bool(message))
                return JsonResponse({
                    'status': 'error',
                    'message': 'Agent IDs and message are required'
                }, status=400)

            notification_sent = False
            failed_agents = []
            
            for agent_id in agent_ids:
                print(f"Processing agent ID: {agent_id}")
                try:
                    agent = AgentUser.objects.get(id=agent_id)
                    print(f"Found agent: {agent.id}")
                    notification = AgentNotification.objects.create(
                        agent_user=agent,
                        message=message
                    )
                    print(f"Created notification: {notification.id}")
                    notification_sent = True
                    logger.info(f"Notification sent to agent {agent.id}")
                except AgentUser.DoesNotExist:
                    print(f"Agent {agent_id} not found in database")
                    failed_agents.append(agent_id)
                    logger.error(f"Agent {agent_id} not found")
                except Exception as e:
                    print(f"Error sending notification to {agent_id}: {str(e)}")
                    failed_agents.append(agent_id)
                    logger.error(f"Failed to send notification to agent {agent_id}: {str(e)}")

            if not notification_sent:
                print("No notifications were sent successfully")
                return JsonResponse({
                    'status': 'error',
                    'message': 'Failed to send notifications to any agents',
                    'failed_agents': failed_agents
                }, status=500)

            response_data = {
                'status': 'success',
                'message': 'Notifications sent successfully'
            }
            
            if failed_agents:
                print(f"Some notifications failed. Failed agents: {failed_agents}")
                response_data['warning'] = 'Some notifications failed to send'
                response_data['failed_agents'] = failed_agents

            print("Returning response:", response_data)
            return JsonResponse(response_data)

        except json.JSONDecodeError:
            print("Failed to decode JSON data")
            logger.error("Invalid JSON data received")
            return JsonResponse({
                'status': 'error',
                'message': 'Invalid JSON data'
            }, status=400)
        except Exception as e:
            print(f"Unexpected error: {str(e)}")
            logger.error(f"Error occurred: {str(e)}", exc_info=True)
            return JsonResponse({
                'status': 'error', 
                'message': 'An error occurred while sending notifications'
            }, status=500)
        
class SendEmployeeNotificationView(View):
    def post(self, request):
        try:
            print("Received request body:", request.body)
            data = json.loads(request.body)
            employee_ids = data.get('employee_ids', [])
            message = data.get('message')
            print("Parsed data:", data)
            print("Employee IDs:", employee_ids)
            print("Message:", message)

            if not employee_ids or not message:
                print("Missing required fields - employees:", bool(employee_ids), "message:", bool(message))
                return JsonResponse({
                    'status': 'error',
                    'message': 'Employee IDs and message are required'
                }, status=400)

            notification_sent = False
            failed_employees = []
            
            for employee_id in employee_ids:
                print(f"Processing employee ID: {employee_id}")
                try:
                    employee = Employee.objects.get(id=employee_id)
                    print(f"Found employee: {employee.id}")
                    notification = EmployeeNotification.objects.create(
                        employee_user=employee,
                        message=message
                    )
                    print(f"Created notification: {notification.id}")
                    notification_sent = True
                    logger.info(f"Notification sent to employee {employee.id}")
                except Employee.DoesNotExist:
                    print(f"Employee {employee_id} not found in database")
                    failed_employees.append(employee_id)
                    logger.error(f"Employee {employee_id} not found")
                except Exception as e:
                    print(f"Error sending notification to {employee_id}: {str(e)}")
                    failed_employees.append(employee_id)
                    logger.error(f"Failed to send notification to employee {employee_id}: {str(e)}")

            if not notification_sent:
                print("No notifications were sent successfully")
                return JsonResponse({
                    'status': 'error',
                    'message': 'Failed to send notifications to any employees',
                    'failed_employees': failed_employees
                }, status=500)

            response_data = {
                'status': 'success',
                'message': 'Notifications sent successfully'
            }
            
            if failed_employees:
                print(f"Some notifications failed. Failed employees: {failed_employees}")
                response_data['warning'] = 'Some notifications failed to send'
                response_data['failed_employees'] = failed_employees

            print("Returning response:", response_data)
            return JsonResponse(response_data)

        except json.JSONDecodeError:
            print("Failed to decode JSON data")
            logger.error("Invalid JSON data received")
            return JsonResponse({
                'status': 'error',
                'message': 'Invalid JSON data'
            }, status=400)
        except Exception as e:
            print(f"Unexpected error: {str(e)}")
            logger.error(f"Error occurred: {str(e)}", exc_info=True)
            return JsonResponse({
                'status': 'error', 
                'message': 'An error occurred while sending notifications'
            }, status=500)        
class SendSupportNotificationView(View):
    def post(self, request):
        try:
            print("Received request body:", request.body)
            data = json.loads(request.body)
            support_ids = data.get('support_ids', []) if isinstance(data.get('support_ids'), list) else [data.get('support_ids')] if data.get('support_ids') else []
            message = data.get('message')
            print("Parsed data:", data)
            print("Support IDs:", support_ids)
            print("Message:", message)

            if not support_ids or not message:
                print("Missing required fields - support:", bool(support_ids), "message:", bool(message))
                return JsonResponse({
                    'status': 'error',
                    'message': 'Support IDs and message are required'
                }, status=400)

            notification_sent = False
            failed_supports = []
            
            for support_id in support_ids:
                print(f"Processing support ID: {support_id}")
                try:
                    support = SupportUser.objects.get(id=support_id)
                    print(f"Found support: {support.id}")
                    notification = SupportNotification.objects.create(
                        support_user=support,
                        message=message
                    )
                    print(f"Created notification: {notification.id}")
                    notification_sent = True
                    logger.info(f"Notification sent to support {support.id}")
                except SupportUser.DoesNotExist:
                    print(f"Support {support_id} not found in database")
                    failed_supports.append(support_id)
                    logger.error(f"Support {support_id} not found")
                except Exception as e:
                    print(f"Error sending notification to {support_id}: {str(e)}")
                    failed_supports.append(support_id)
                    logger.error(f"Failed to send notification to support {support_id}: {str(e)}")

            if not notification_sent:
                print("No notifications were sent successfully")
                return JsonResponse({
                    'status': 'error',
                    'message': 'Failed to send notifications to any supports',
                    'failed_supports': failed_supports
                }, status=500)

            response_data = {
                'status': 'success',
                'message': 'Notifications sent successfully'
            }
            
            if failed_supports:
                print(f"Some notifications failed. Failed supports: {failed_supports}")
                response_data['warning'] = 'Some notifications failed to send'
                response_data['failed_supports'] = failed_supports

            print("Returning response:", response_data)
            return JsonResponse(response_data)

        except json.JSONDecodeError:
            print("Failed to decode JSON data")
            logger.error("Invalid JSON data received")
            return JsonResponse({
                'status': 'error',
                'message': 'Invalid JSON data'
            }, status=400)
        except Exception as e:
            print(f"Unexpected error: {str(e)}")
            logger.error(f"Error occurred: {str(e)}", exc_info=True)
            return JsonResponse({
                'status': 'error', 
                'message': 'An error occurred while sending notifications'
            }, status=500)
        

class UserDetailsView(TemplateView):
    template_name = 'masteradmin/user_details.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['total_agents'] = AgentUser.objects.filter(name__isnull=False, is_approved=False).count() + Employee.objects.filter(name__isnull=False, is_approved=False).count() + SupportUser.objects.filter(name__isnull=False, is_approved=False).count()
        context['total_meetings'] = Meeting.objects.filter(is_active=True).count()
        context['departments'] = Policy.objects.all()
        context['support_departments'] = SupportDepartment.objects.all()
        return context
    
def approve_employee(request, employee_id):
    try:
        employee = Employee.objects.get(id=employee_id)
        employee.is_approved = True
        employee.is_active = True
        employee.is_rejected = False
        employee.save()

        # Send approval email
        try:
            subject = 'Your Account Has Been Approved'
            message = f"""
            Dear {employee.name},

            Your account has been approved. You can now login to your account using your registered email and password.

            Login here: https://{request.get_host()}/employee/

            Thank you for joining us!

            Best regards,
            The Admin Team
            """
            send_mail(
                subject,
                message,
                settings.DEFAULT_FROM_EMAIL,
                [employee.email],
                fail_silently=False,
            )
            logger.info(f"Approval email sent successfully to employee {employee.email}")
        except Exception as e:
            logger.error(f"Failed to send approval email to {employee.email}: {str(e)}")
            # Continue execution even if email fails

        return redirect('agent_details')
    except Employee.DoesNotExist:
        messages.error(request, 'Employee not found')
        return redirect('agent_details')
    except Exception as e:
        logger.error(f"Error approving employee: {str(e)}")
        messages.error(request, 'Error approving employee')
        return redirect('agent_details')

def reject_employee(request, employee_id):
    try:
        employee = Employee.objects.get(id=employee_id)
        employee.is_rejected = True 
        employee.is_active = False
        employee.save()

        # Send rejection email
        try:
            subject = 'Your Account Has Been Rejected'
            message = f"""
            Dear {employee.name},

            We regret to inform you that your account has been rejected.

            Thank you for your interest in joining us.

            Best regards,
            The Admin Team
            """
            send_mail(
                subject,
                message,
                settings.DEFAULT_FROM_EMAIL,
                [employee.email],
                fail_silently=False,
            )
            logger.info(f"Rejection email sent successfully to employee {employee.email}")
        except Exception as e:
            logger.error(f"Failed to send rejection email to {employee.email}: {str(e)}")
            # Continue execution even if email fails

        return redirect('agent_details')
    except Employee.DoesNotExist:
        messages.error(request, 'Employee not found')
        return redirect('agent_details')
    except Exception as e:
        logger.error(f"Error rejecting employee: {str(e)}")
        messages.error(request, 'Error rejecting employee')
        return redirect('agent_details')

def approve_support(request, support_id):
    try:
        support = SupportUser.objects.get(id=support_id)
        if support.is_approved:
            return JsonResponse({'status': 'error', 'message': 'Support is already approved'}, status=400)
        
        support.is_approved = True
        support.is_active = True
        support.is_rejected = False
        support.save()

        # Send approval email to support
        try:
            subject = 'Your Account Has Been Approved'
            message = f"""
            Dear {support.name},

            Your account has been approved. You can now login to your account using your registered email and password.

            Thank you for joining us!

            login : https://{request.get_host()}/customersupport/

            Best regards,
            The Admin Team
            """
            send_mail(
                subject,
                message,
                settings.DEFAULT_FROM_EMAIL,
                [support.email],
                fail_silently=False,
            )
            logger.info(f"Approval email sent successfully to support {support.email}")
        except Exception as e:
            logger.error(f"Failed to send approval email to {support.email}: {str(e)}")
            # Continue execution even if email fails
            
        return redirect('support_details')
    except SupportUser.DoesNotExist:
        messages.error(request, 'Support user not found')
        return redirect('support_details')
    except Exception as e:
        logger.error(f"Error approving support: {str(e)}")
        messages.error(request, 'Error approving support')
        return redirect('support_details')

def reject_support(request, support_id):
    try:
        support = SupportUser.objects.get(id=support_id)
        support.is_rejected = True
        support.is_active = False
        support.save()

        # Send rejection email
        try:
            subject = 'Your Account Has Been Rejected'
            message = f"""
            Dear {support.name},

            We regret to inform you that your account has been rejected.

            Thank you for your interest in joining us.

            Best regards,
            The Admin Team
            """
            send_mail(
                subject,
                message,
                settings.DEFAULT_FROM_EMAIL,
                [support.email],
                fail_silently=False,
            )
            logger.info(f"Rejection email sent successfully to support {support.email}")
        except Exception as e:
            logger.error(f"Failed to send rejection email to {support.email}: {str(e)}")
            # Continue execution even if email fails

        return redirect('support_details')
    except SupportUser.DoesNotExist:
        messages.error(request, 'Support user not found')
        return redirect('support_details')
    except Exception as e:
        logger.error(f"Error rejecting support: {str(e)}")
        messages.error(request, 'Error rejecting support')
        return redirect('support_details')


class DeleteSupportDepartmentView(View):
    def post(self, request, department_id, *args, **kwargs):
        try:
            department = SupportDepartment.objects.get(id=department_id)
            department.delete()
            return JsonResponse({
                'status': 'success',
                'message': 'Department deleted successfully'
            })
        except SupportDepartment.DoesNotExist:
            return JsonResponse({
                'status': 'error',
                'message': 'Department not found'
            }, status=404)
        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'message': str(e)
            }, status=500)


class AI_PromptView(TemplateView):
    template_name = 'masteradmin/ai_prompt.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['ai_prompts'] = AI_Prompt.objects.all()
        context['total_agents'] = AgentUser.objects.filter(name__isnull=False, is_approved=False).count() + Employee.objects.filter(name__isnull=False, is_approved=False).count() + SupportUser.objects.filter(name__isnull=False, is_approved=False).count()
        return context

    def post(self, request, *args, **kwargs):
        try:
            data = json.loads(request.body)
            prompt_text = data.get('prompt', '').strip()
            
            # Split prompt into lines
            lines = prompt_text.split('\n')
            if len(lines) < 3:  # Need at least action identifier, prompt content, and close identifier
                return JsonResponse({
                    'status': 'error',
                    'message': 'Invalid prompt format. Must include action identifier, prompt content, and close identifier.'
                }, status=400)

            action_identifier = lines[0].strip()
            prompt_content = '\n'.join(lines[1:-1]).strip()
            close_identifier = lines[-1].strip()

            # Validate close identifier
            # Extract platform name from action identifier
            platform_name = action_identifier[1:].strip() if action_identifier.startswith('#') else action_identifier[2:].strip()
            expected_close = f'Done {platform_name}'

            if close_identifier != expected_close:
                return JsonResponse({
                    'status': 'error', 
                    'message': f'Invalid close identifier'
                }, status=400)
            # Handle new prompt creation
            if action_identifier.startswith('#'):
                platform = action_identifier[1:].strip()  # Remove # and whitespace
                if not platform or not prompt_content:
                    return JsonResponse({
                        'status': 'error',
                        'message': 'Platform name and prompt content are required'
                    }, status=400)

                # Create new AI prompt with only the prompt content
                AI_Prompt.objects.create(
                    platform=platform,
                    prompt=prompt_content
                )
                return JsonResponse({
                    'status': 'success',
                    'message': f'Created new prompt for platform: {platform}'
                })

            # Handle prompt update
            elif action_identifier.startswith('e-'):
                platform = action_identifier[2:].strip()  # Remove e- and whitespace
                if not platform or not prompt_content:
                    return JsonResponse({
                        'status': 'error',
                        'message': 'Platform name and prompt content are required'
                    }, status=400)

                # Find and update existing prompt with only the prompt content
                try:
                    existing_prompt = AI_Prompt.objects.get(platform=platform)
                    existing_prompt.prompt = prompt_content
                    existing_prompt.save()
                    return JsonResponse({
                        'status': 'success',
                        'message': f'Updated prompt for platform: {platform}'
                    })
                except AI_Prompt.DoesNotExist:
                    return JsonResponse({
                        'status': 'error',
                        'message': f'No existing prompt found for platform: {platform}'
                    }, status=404)

            else:
                return JsonResponse({
                    'status': 'error',
                    'message': 'Invalid action identifier. Must start with # or e-'
                }, status=400)

        except json.JSONDecodeError:
            return JsonResponse({
                'status': 'error',
                'message': 'Invalid JSON data'
            }, status=400)
        except Exception as e:
            logger.error(f"Error processing AI prompt: {str(e)}", exc_info=True)
            return JsonResponse({
                'status': 'error',
                'message': f'Error processing AI prompt: {str(e)}'
            }, status=500)

class CreateTicketView(View):
    def post(self, request):
        try:
            data = json.loads(request.body)
            
            # Extract data from request
            title = data.get('title')
            description = data.get('description')
            priority = data.get('priority')
            support_user_id = data.get('support_user')
            
            # Validate required fields
            if not all([title, description, priority]):
                return JsonResponse({
                    'status': 'error',
                    'message': 'Title, description, and priority are required fields'
                }, status=400)

            # Validate priority values
            valid_priorities = ['low', 'medium', 'high', 'urgent']
            if priority not in valid_priorities:
                return JsonResponse({
                    'status': 'error',
                    'message': 'Invalid priority level'
                }, status=400)

            # Get support user if specified
            support_user = None
            if support_user_id:
                try:
                    support_user = SupportUser.objects.get(id=support_user_id)
                except SupportUser.DoesNotExist:
                    return JsonResponse({
                        'status': 'error',
                        'message': 'Selected support user not found'
                    }, status=404)

            # Create ticket
            ticket = Tickets.objects.create(
                title=title,
                description=description,
                priority=priority,
                status='open',
                assigned_to=support_user,
                created_by=request.user
            )

            # Send notification to assigned support user if any
            if support_user:
                try:
                    notification = SupportNotification.objects.create(
                        support_user=support_user,
                        message=f'New ticket assigned: {title}',
                        ticket=ticket
                    )
                    logger.info(f"Notification sent to support user {support_user.id}")
                except Exception as e:
                    logger.error(f"Failed to send notification: {str(e)}")

            return JsonResponse({
                'status': 'success',
                'message': 'Ticket created successfully',
                'ticket': {
                    'id': ticket.id,
                    'title': ticket.title,
                    'priority': ticket.priority,
                    'status': ticket.status,
                    'assigned_to': support_user.name if support_user else None
                }
            })

        except json.JSONDecodeError:
            return JsonResponse({
                'status': 'error',
                'message': 'Invalid JSON data'
            }, status=400)
        except Exception as e:
            logger.error(f"Error creating ticket: {str(e)}", exc_info=True)
            return JsonResponse({
                'status': 'error',
                'message': 'An error occurred while creating the ticket'
            }, status=500)

class MarkAgentAsInactiveView(View):
    def post(self, request, *args, **kwargs):
        agent_id = self.kwargs.get('agent_id')
        try:
            agent = AgentUser.objects.get(id=agent_id)
            if not agent.is_active:
                return JsonResponse({'status': 'error', 'message': 'Agent is already inactive'}, status=400)
            
            agent.is_active = False
            agent.save()

            # Send deactivation email to agent
            try:
                subject = 'Your Account Has Been Deactivated'
                message = f"""
                Dear {agent.name},

                Your account has been deactivated. You will no longer be able to login until your account is reactivated.

                If you have any questions, please contact the admin team.

                Best regards,
                The Admin Team
                """
                send_mail(
                    subject,
                    message,
                    settings.DEFAULT_FROM_EMAIL,
                    [agent.email],
                    fail_silently=False,
                )
                logger.info(f"Deactivation email sent successfully to agent {agent.email}")
            except Exception as e:
                logger.error(f"Failed to send deactivation email to {agent.email}: {str(e)}")
                # Continue execution even if email fails
                
            return JsonResponse({'status': 'success', 'message': 'Agent deactivated successfully'})
        except AgentUser.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Agent not found'}, status=404)

class MarkAgentAsActiveView(View):
    def post(self, request, *args, **kwargs):
        agent_id = self.kwargs.get('agent_id')
        try:
            agent = AgentUser.objects.get(id=agent_id)
            if agent.is_active:
                return JsonResponse({'status': 'error', 'message': 'Agent is already active'}, status=400)
            
            agent.is_active = True
            agent.save()

            # Send activation email to agent
            try:
                subject = 'Your Account Has Been Activated'
                message = f"""
                Dear {agent.name},

                Your account has been reactivated. You can now login to your account.

                If you have any questions, please contact the admin team.

                Best regards,
                The Admin Team
                """
                send_mail(
                    subject,
                    message,
                    settings.DEFAULT_FROM_EMAIL,
                    [agent.email],
                    fail_silently=False,
                )
                logger.info(f"Activation email sent successfully to agent {agent.email}")
            except Exception as e:
                logger.error(f"Failed to send activation email to {agent.email}: {str(e)}")
                # Continue execution even if email fails
                
            return JsonResponse({'status': 'success', 'message': 'Agent activated successfully'})
        except AgentUser.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Agent not found'}, status=404)

class TotalSubscriptionView(TemplateView):
    template_name = 'masteradmin/total_subscription.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['total_subscriptions'] = Subscription.objects.count()
        return context
    
class SuspendSubscriptionView(View):
    def post(self, request, *args, **kwargs):
        subscription_id = self.kwargs.get('subscription_id')
        try:
            subscription = Subscription.objects.get(id=subscription_id)
            # Toggle both suspended and active status
            subscription.is_suspended = not subscription.is_suspended
            subscription.is_active = not subscription.is_suspended  # Active should be opposite of suspended
            subscription.save()
            
            status = 'suspended' if subscription.is_suspended else 'activated'
            return JsonResponse({
                'status': 'success',
                'message': f'Subscription {status} successfully',
                'is_suspended': subscription.is_suspended
            })
        except Subscription.DoesNotExist:
            return JsonResponse({
                'status': 'error',
                'message': 'Subscription not found'
            }, status=404)
        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'message': str(e)
            }, status=500)

class CreateUserPolicy(TemplateView):
    template_name = 'masteradmin/user_policy.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['user_policies'] = UserPolicy.objects.all()
        return context

    def post(self, request, *args, **kwargs):
        data = json.loads(request.body)
        name = data.get('name')
        description = data.get('description')
        UserPolicy.objects.create(name=name, description=description)
        return JsonResponse({'status': 'success', 'message': 'User policy created successfully'})
class CreateUserNotification(View):
    def post(self, request, *args, **kwargs):
        try:
            data = json.loads(request.body)
            title = data.get('title')
            message = data.get('message')
            link = data.get('link')
            attachment = data.get('attachment')
            user_ids = data.get('users', [])

            # Validate required fields
            if not all([title, message]):
                return JsonResponse({
                    'status': 'error', 
                    'message': 'Title and message are required fields'
                }, status=400)

            # Create notification without users first
            notification = UserNotification.objects.create(
                title=title,
                message=message, 
                link=link,
                attachment=attachment
            )

            # Add users through both the through model and direct users field
            if user_ids:
                users = User.objects.filter(id__in=user_ids)
                notification.users.add(*users)  # Add to users field
                for user in users:
                    UserNotificationRecipient.objects.create(
                        notification=notification,
                        user=user
                    )

            return JsonResponse({
                'status': 'success',
                'message': 'User notification created successfully',
                'notification_id': str(notification.id)
            })

        except json.JSONDecodeError:
            return JsonResponse({
                'status': 'error',
                'message': 'Invalid JSON data'
            }, status=400)
            
        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'message': str(e)
            }, status=500)

def get_feedbacks(request):
    page = request.GET.get('page', 1)
    feedback_type = request.GET.get('type', 'all')  # New parameter for filter type
    
    try:
        page = int(page)
    except ValueError:
        page = 1
    
    # Base queryset
    feedbacks = Feedbacks.objects.all()
    
    # Apply filters based on feedback type
    if feedback_type == 'positive':
        feedbacks = feedbacks.filter(rating__in=[4, 5])
    elif feedback_type == 'negative':
        feedbacks = feedbacks.filter(rating__in=[2, 3])
    elif feedback_type == 'one_star':
        feedbacks = feedbacks.filter(rating=1)
    
    # Order by created_at
    feedbacks = feedbacks.order_by('-created_at')
    
    paginator = Paginator(feedbacks, 6)  # Show 6 feedbacks per page
    
    try:
        feedbacks_page = paginator.page(page)
    except (EmptyPage, PageNotAnInteger):
        feedbacks_page = paginator.page(1)
    
    feedbacks_data = [{
        'name': feedback.name,
        'rating': feedback.rating,
        'message': feedback.message,
        'created_at': feedback.created_at.isoformat()
    } for feedback in feedbacks_page]
    
    return JsonResponse({
        'feedbacks': feedbacks_data,
        'has_more': feedbacks_page.has_next(),
        'total_pages': paginator.num_pages,
        'current_page': page,
        'type': feedback_type  # Return the feedback type
    })

class DeleteUserPolicy(View):
    def post(self, request, *args, **kwargs):
        try:
            policy_id = self.kwargs.get('policy_id')
            UserPolicy.objects.get(id=policy_id).delete()
            return JsonResponse({'status': 'success', 'message': 'User policy deleted successfully'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    
    
def suspend_support(request, support_id):
    support = SupportUser.objects.get(id=support_id)
    support.is_suspended = True
    support.is_active = False
    support.is_approved = False
    support.is_rejected = False
    support.save()
    return redirect('mastersupport')

def mark_support_as_active(request, support_id):
    support = SupportUser.objects.get(id=support_id)
    support.is_suspended = False
    support.is_active = True
    support.is_approved = True
    support.is_rejected = False
    support.save()
    return redirect('mastersupport')

def reject_support(request, support_id):
    support = SupportUser.objects.get(id=support_id)
    support.is_suspended = False
    support.is_active = False
    support.is_approved = False
    support.is_rejected = True
    support.save()
    return redirect('mastersupport')

@csrf_protect
@require_http_methods(["POST"])
def update_support_salary(request):
    try:
        data = json.loads(request.body)
        support_id = data.get('support_id')
        salary = data.get('salary')

        if not support_id or salary is None:
            return JsonResponse({'status': 'error', 'message': 'Missing required fields'}, status=400)

        support_user = SupportUser.objects.get(id=support_id)
        support_user.salary = Decimal(salary) if salary else None
        support_user.save()

        return JsonResponse({'status': 'success', 'message': 'Salary updated successfully'})
    except SupportUser.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Support user not found'}, status=404)
    except (ValueError, json.JSONDecodeError):
        return JsonResponse({'status': 'error', 'message': 'Invalid data provided'}, status=400)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    
class AssignToMeView(View):
    def get(self, request, ticket_id):
        try:
            ticket = Tickets.objects.get(id=ticket_id)
            # Get the current support user (you'll need to implement this based on your authentication)
            current_user = None  # Adjust this based on your user model
            
            ticket.assigned_to = current_user
            ticket.save()

            # Redirect back to the ticket list
            return redirect('ticket')  # Adjust the URL name as needed

        except Tickets.DoesNotExist:
            messages.error(request, 'Ticket not found')
            return redirect('ticket')
        except Exception as e:
            messages.error(request, str(e))
            return redirect('ticket')

    def post(self, request, ticket_id):
        try:
            ticket = Tickets.objects.get(id=ticket_id)
            # Get the current support user
            current_user = None # Adjust this based on your user model
            
            ticket.assigned_to = current_user
            ticket.save()

            return JsonResponse({
                'status': 'success',
                'message': 'Ticket assigned to you successfully'
            })

        except Tickets.DoesNotExist:
            return JsonResponse({
                'status': 'error',
                'message': 'Ticket not found'
            }, status=404)
        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'message': str(e)
            }, status=500)

class GetSupportUsersView(View):
    def get(self, request, department_id):
        try:
            # Get active support users from the department
            support_users = SupportUser.objects.filter(
                support_department_id=department_id,
                is_active=True,
                is_approved=True,
                is_rejected=False,
                is_suspended=False
            ).annotate(
                pending_tickets=Count('tickets', filter=Q(tickets__status='Pending'))
            )

            users_data = [{
                'id': str(user.id),
                'name': user.name,
                'pending_tickets': user.pending_tickets
            } for user in support_users]

            return JsonResponse({
                'status': 'success',
                'users': users_data
            })
        except Exception as e:
            logger.exception("Error getting support users")
            return JsonResponse({
                'status': 'error',
                'message': str(e)
            }, status=500)

class AssignTicketView(View):
    def post(self, request, ticket_id):
        logger.info(f"Assigning ticket {ticket_id}")
        try:
            data = json.loads(request.body)
            support_user_id = data.get('support_user_id')

            if not support_user_id:
                return JsonResponse({
                    'status': 'error',
                    'message': 'Support user ID is required'
                }, status=400)

            try:
                ticket = Tickets.objects.get(id=ticket_id)
            except Tickets.DoesNotExist:
                return JsonResponse({
                    'status': 'error',
                    'message': 'Ticket not found'
                }, status=404)

            try:
                support_user = SupportUser.objects.get(id=support_user_id)
            except SupportUser.DoesNotExist:
                return JsonResponse({
                    'status': 'error',
                    'message': 'Support user not found'
                }, status=404)

            # Update the ticket assignment
            ticket.assigned_to = support_user
            ticket.save()

            logger.info(f"Successfully assigned ticket {ticket_id} to {support_user.name}")
            return JsonResponse({
                'status': 'success',
                'message': f'Ticket successfully assigned to {support_user.name}'
            })

        except json.JSONDecodeError:
            return JsonResponse({
                'status': 'error',
                'message': 'Invalid JSON data'
            }, status=400)
        except Exception as e:
            logger.exception("Error assigning ticket")
            return JsonResponse({
                'status': 'error',
                'message': str(e)
            }, status=500)

class DebugTicketAssignmentView(View):
    def get(self, request, ticket_id):
        try:
            ticket = Tickets.objects.get(id=ticket_id)
            return JsonResponse({
                'ticket_id': str(ticket.id),
                'assigned_to': str(ticket.assigned_to.id) if ticket.assigned_to else None,
                'status': ticket.status
            })
        except Tickets.DoesNotExist:
            return JsonResponse({
                'error': 'Ticket not found'
            }, status=404)

class MyTicketsView(TemplateView):
    template_name = 'masteradmin/my_tickets.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        masteradmin_id = self.request.session.get('masteradmin_id')
        if masteradmin_id:
            tickets = Tickets.objects.filter(assigned_to=masteradmin_id)
            context['tickets'] = tickets
        return context


class BeesuggestView(TemplateView):
    """
    View to display all product submissions for masteradmin review
    """
    template_name = 'masteradmin/beesuggest.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get all product submissions with pagination
        products_list = ProductDetails.objects.all().order_by('-submitted_at')
        
        # Filter by status if provided
        status_filter = self.request.GET.get('status')
        if status_filter == 'published':
            products_list = products_list.filter(is_published=True)
        elif status_filter == 'pending':
            products_list = products_list.filter(is_published=False)
        
        # Search functionality
        search_query = self.request.GET.get('search')
        if search_query:
            products_list = products_list.filter(
                Q(organization__icontains=search_query) |
                Q(focus_keywords__icontains=search_query) |
                Q(product_description__icontains=search_query) |
                Q(user__username__icontains=search_query) |
                Q(user__email__icontains=search_query)
            )
        
        # Pagination
        paginator = Paginator(products_list, 10)  # Show 10 products per page
        page_number = self.request.GET.get('page')
        products = paginator.get_page(page_number)
        
        # Statistics
        total_products = ProductDetails.objects.count()
        published_products = ProductDetails.objects.filter(is_published=True).count()
        pending_products = ProductDetails.objects.filter(is_published=False).count()
        
        context.update({
            'products': products,
            'total_products': total_products,
            'published_products': published_products,
            'pending_products': pending_products,
            'search_query': search_query,
            'status_filter': status_filter,
        })
        
        return context


class EditProductView(View):
    """
    View to edit and manage individual product submissions
    """
    
    def get(self, request, product_id):
        """Display the edit form for a product"""
        try:
            product = ProductDetails.objects.get(id=product_id)
            context = {
                'product': product,
                'variations': product.variations if product.variations else [],
                'faqs': product.faqs if product.faqs else [],
            }
            return render(request, 'masteradmin/edit_product.html', context)
        except ProductDetails.DoesNotExist:
            messages.error(request, 'Product not found.')
            return redirect('beesuggest')
    
    def post(self, request, product_id):
        """Handle product updates"""
        try:
            product = ProductDetails.objects.get(id=product_id)
            
            # Handle publish/unpublish action
            action = request.POST.get('action')
            if action == 'publish':
                product.is_published = True
                product.published_at = timezone.now()
                product.save()
                messages.success(request, f'Product "{product.organization or "Unnamed"}" has been published successfully.')
                return redirect('beesuggest')
            elif action == 'unpublish':
                product.is_published = False
                product.published_at = None
                product.save()
                messages.success(request, f'Product "{product.organization or "Unnamed"}" has been unpublished.')
                return redirect('beesuggest')
            elif action == 'delete':
                product_name = product.organization or "Unnamed"
                product.delete()
                messages.success(request, f'Product "{product_name}" has been deleted successfully.')
                return redirect('beesuggest')
            
            # Handle form updates
            # Basic fields
            product.focus_keywords = request.POST.get('focus_keywords', '')
            product.alt_keyword_1 = request.POST.get('alt_keyword_1', '')
            product.alt_keyword_2 = request.POST.get('alt_keyword_2', '')
            
            # Handle file uploads if new files are provided
            for i in range(1, 6):
                image_field = f'product_image_{i}'
                alt_field = f'product_image_{i}_alt'
                if image_field in request.FILES:
                    setattr(product, image_field, request.FILES[image_field])
                setattr(product, alt_field, request.POST.get(alt_field, ''))
            
            product.video_url_1 = request.POST.get('video_url_1', '')
            product.video_url_2 = request.POST.get('video_url_2', '')
            product.video_url_3 = request.POST.get('video_url_3', '')
            
            if 'size_chart' in request.FILES:
                product.size_chart = request.FILES['size_chart']
            
            # Text fields
            product.product_description = request.POST.get('product_description', '')
            product.uses = request.POST.get('uses', '')
            product.best_suited_for = request.POST.get('best_suited_for', '')
            product.social_media_facebook = request.POST.get('social_media_facebook', '')
            product.social_media_twitter = request.POST.get('social_media_twitter', '')
            product.social_media_instagram = request.POST.get('social_media_instagram', '')
            product.website_url = request.POST.get('website_url', '')
            product.contact_number = request.POST.get('contact_number', '')
            product.email = request.POST.get('email', '')
            product.address = request.POST.get('address', '')
            product.organization = request.POST.get('organization', '')
            product.gst_details = request.POST.get('gst_details', '')
            product.map_location = request.POST.get('map_location', '')
            product.why_choose_us = request.POST.get('why_choose_us', '')
            product.comparison = request.POST.get('comparison', '')

            # Process variations
            variations = []
            variation_count = 1
            while f'variation_name_{variation_count}' in request.POST:
                name = request.POST.get(f'variation_name_{variation_count}')
                value = request.POST.get(f'variation_value_{variation_count}')
                if name and value:
                    variations.append({'name': name, 'value': value})
                variation_count += 1
            product.variations = variations

            # Process FAQs
            faqs = []
            faq_count = 1
            while f'faq_question_{faq_count}' in request.POST:
                question = request.POST.get(f'faq_question_{faq_count}')
                answer = request.POST.get(f'faq_answer_{faq_count}')
                if question and answer:
                    faqs.append({'question': question, 'answer': answer})
                faq_count += 1
            product.faqs = faqs
            
            # Save the product
            product.save()
            
            messages.success(request, f'Product "{product.organization or "Unnamed"}" has been updated successfully.')
            return redirect('edit_product', product_id=product_id)
            
        except ProductDetails.DoesNotExist:
            messages.error(request, 'Product not found.')
            return redirect('beesuggest')
        except Exception as e:
            logger.error(f"Error updating product: {str(e)}")
            messages.error(request, 'There was an error updating the product. Please try again.')
            return redirect('edit_product', product_id=product_id)

class BeesuggestAgreementView(View):
    """
    View for masteradmin to create or update the Beesuggest Agreement.
    Ensures that there is only one active agreement.
    """
    template_name = 'masteradmin/beesuggest_agreement.html'

    def get(self, request, *args, **kwargs):
        agreements = BeesuggestAgreement.objects.all()
        active_agreement = BeesuggestAgreement.objects.filter(is_active=True).first()
        context = {
            'agreements': agreements,
            'active_agreement': active_agreement,
            'l_page': 'Beesuggest',
            'i_page': 'Agreement',
        }
        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        title = request.POST.get('title')
        content = request.POST.get('content')

        if not title or not content:
            messages.error(request, 'Title and content cannot be empty.')
            return redirect('beesuggest_agreement')

        # Create a new agreement. The save method on the model will handle is_active.
        # For a new agreement, we might want to decide if it should become active automatically.
        # Let's assume for now new agreements are NOT active by default and must be explicitly activated.
        BeesuggestAgreement.objects.create(
            title=title,
            content=content,
            is_active=False # Default to not active
        )
        
        messages.success(request, 'New agreement has been created successfully. You can now set it as the active agreement.')
        
        return redirect('beesuggest_agreement')

@method_decorator(csrf_exempt, name='dispatch')
class SetActiveBeesuggestAgreementView(View):
    def post(self, request, *args, **kwargs):
        agreement_id = request.POST.get('agreement_id')
        try:
            agreement = BeesuggestAgreement.objects.get(id=agreement_id)
            agreement.is_active = True
            agreement.save()
            messages.success(request, f'Agreement "{agreement.title}" has been set as active.')
        except BeesuggestAgreement.DoesNotExist:
            messages.error(request, 'Agreement not found.')
        except Exception as e:
            messages.error(request, f'An error occurred: {e}')
        return redirect('beesuggest_agreement')

@method_decorator(csrf_exempt, name='dispatch')
class DeleteBeesuggestAgreementView(View):
    def post(self, request, *args, **kwargs):
        agreement_id = request.POST.get('agreement_id')
        try:
            agreement = BeesuggestAgreement.objects.get(id=agreement_id)
            if agreement.is_active:
                messages.error(request, 'Cannot delete an active agreement. Please activate another agreement first.')
            else:
                title = agreement.title
                agreement.delete()
                messages.success(request, f'Agreement "{title}" has been deleted.')
        except BeesuggestAgreement.DoesNotExist:
            messages.error(request, 'Agreement not found.')
        except Exception as e:
            messages.error(request, f'An error occurred: {e}')
        return redirect('beesuggest_agreement')

    def post(self, request, *args, **kwargs):
        try:
            data = json.loads(request.body)
            agreement_id = data.get('agreement_id')
            
            agreement = BeesuggestAgreement.objects.get(id=agreement_id)
            agreement.delete()
            
            return JsonResponse({'status': 'success', 'message': 'Agreement deleted successfully.'})
        
        except BeesuggestAgreement.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Agreement not found.'}, status=404)
        
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


class UserAgreementView(View):
    """
    View for masteradmin to create or update the User Agreement.
    Ensures that there is only one active agreement.
    """
    template_name = 'masteradmin/user_agreement.html'

    def get(self, request, *args, **kwargs):
        agreements = UserAgreement.objects.all().order_by('-created_at')
        active_agreement = UserAgreement.objects.filter(is_active=True).first()
        
        context = {
            'agreements': agreements,
            'active_agreement': active_agreement,
        }
        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        title = request.POST.get('title')
        content = request.POST.get('content')
        
        if not title or not content:
            messages.error(request, 'Title and content are required.')
            return redirect('user_agreement')

        # Create new agreement
        new_agreement = UserAgreement.objects.create(title=title, content=content)
        
        # Set all other agreements to inactive
        UserAgreement.objects.exclude(id=new_agreement.id).update(is_active=False)
        
        # Set the new agreement to active
        new_agreement.is_active = True
        new_agreement.save()
        
        messages.success(request, 'New user agreement created and set as active.')
        return redirect('user_agreement')

@method_decorator(csrf_exempt, name='dispatch')
class SetActiveUserAgreementView(View):
    def post(self, request, *args, **kwargs):
        try:
            data = json.loads(request.body)
            agreement_id = data.get('agreement_id')
            
            # Deactivate all other agreements
            UserAgreement.objects.all().update(is_active=False)
            
            # Activate the selected agreement
            agreement = UserAgreement.objects.get(id=agreement_id)
            agreement.is_active = True
            agreement.save()
            
            return JsonResponse({'status': 'success', 'message': 'Agreement set as active.'})
        
        except UserAgreement.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Agreement not found.'}, status=404)
        
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

@method_decorator(csrf_exempt, name='dispatch')
class DeleteUserAgreementView(View):
    def post(self, request, *args, **kwargs):
        try:
            data = json.loads(request.body)
            agreement_id = data.get('agreement_id')
            
            agreement = UserAgreement.objects.get(id=agreement_id)
            agreement.delete()
            
            return JsonResponse({'status': 'success', 'message': 'Agreement deleted successfully.'})
        
        except UserAgreement.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Agreement not found.'}, status=404)
        
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
